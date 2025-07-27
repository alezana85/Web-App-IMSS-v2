import polars as pl
import pandas as pd
import os
import re
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook

def confronta_entre_suas(sua1_path, sua2_path, archive_path):
    sua1_filename = os.path.basename(sua1_path)
    sua2_filename = os.path.basename(sua2_path)
    
    sua1_match = re.search(r'(\d{2})-(\d{4})', sua1_filename)
    sua2_match = re.search(r'(\d{2})-(\d{4})', sua2_filename)
    
    if not sua1_match or not sua2_match:
        raise ValueError("Los nombres de archivo deben contener el formato MM-YYYY")
    
    sua1_mes, sua1_año = sua1_match.groups()
    sua2_mes, sua2_año = sua2_match.groups()
    
    mes = int(sua1_mes)
    año = int(sua1_año)
    
    num_hojas = 2 if mes % 2 == 0 else 1
    
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)
    
    purple_fill = PatternFill(start_color="8B008B", end_color="8B008B", fill_type="solid")
    green_font = Font(color="00FF00", bold=True)
    
    hoja1_result = procesar_hoja_sua_mensual(sua1_path, sua2_path)
    ws1 = wb.create_sheet("SUA_MENSUAL")
    escribir_dataframe_a_excel(ws1, hoja1_result, purple_fill, green_font)
    
    if num_hojas == 2:
        hoja2_result = procesar_hoja_sua_bimestral(sua1_path, sua2_path)
        ws2 = wb.create_sheet("SUA_BIMESTRAL")
        escribir_dataframe_a_excel(ws2, hoja2_result, purple_fill, green_font)
    
    output_filename = f"{sua1_mes}_{sua1_año}_CONFRONTA_SUAS.xlsx"
    output_path = os.path.join(archive_path, output_filename)
    wb.save(output_path)
    
    return output_path

def procesar_hoja_sua_mensual(sua1_path, sua2_path):
    try:
        sua1_df = pl.read_excel(sua1_path, sheet_name="SUA_MENSUAL")
        sua2_df = pl.read_excel(sua2_path, sheet_name="SUA_MENSUAL")
    except Exception as e:
        raise ValueError(f"Error al leer los archivos Excel: {e}")
    
    sua1_df = sua1_df.with_columns([
        pl.concat_str([pl.col("RP").cast(pl.Utf8).str.slice(0, 10), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    sua2_df = sua2_df.with_columns([
        pl.concat_str([pl.col("RP").cast(pl.Utf8).str.slice(0, 10), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    todas_ids = set(sua1_df["ID_UNICO"].to_list()) | set(sua2_df["ID_UNICO"].to_list())
    
    resultados = []
    
    for id_unico in todas_ids:
        sua1_row = sua1_df.filter(pl.col("ID_UNICO") == id_unico)
        sua2_row = sua2_df.filter(pl.col("ID_UNICO") == id_unico)
        
        sua1_existe = len(sua1_row) > 0
        sua2_existe = len(sua2_row) > 0
        
        if sua1_existe:
            sua1_data = sua1_row.to_dicts()[0]
            resultado = sua1_data.copy()
        else:
            sua2_data = sua2_row.to_dicts()[0]
            resultado = {col: None for col in sua1_df.columns}
            resultado["RP"] = sua2_data["RP"]
            resultado["NSS"] = sua2_data["NSS"]
            resultado["ID_UNICO"] = sua2_data["ID_UNICO"]
        
        if sua1_existe and sua2_existe:
            sua1_data = sua1_row.to_dicts()[0]
            sua2_data = sua2_row.to_dicts()[0]
            
            resultado["NOMBRE ASEGURADO"] = sua2_data["NOMBRE ASEGURADO"]
            resultado["RFC"] = sua2_data["RFC"]
            resultado["CURP"] = sua2_data["CURP"]
            resultado["N_MOVS"] = sua2_data["N_MOVS"]
            resultado["SDI"] = sua2_data["SDI"]
            resultado["INC"] = sua1_data.get("INC", 0)
            resultado["AUS"] = sua1_data.get("AUS", 0)
            
            columnas_numericas = ["DIAS", "CF", "EXC_PAT", "EXC_OBR", "PD_PAT", "PD_OBR", 
                                "GMP_PAT", "GMP_OBR", "RT", "IV_PAT", "IV_OBR", "GPS", "TOTAL"]
            
            for columna in columnas_numericas:
                if columna in sua1_data and columna in sua2_data:
                    sua1_val = sua1_data[columna] or 0
                    sua2_val = sua2_data[columna] or 0
                    
                    if isinstance(sua1_val, (float, int)) and isinstance(sua2_val, (float, int)):
                        sua1_val = round(float(sua1_val), 2)
                        sua2_val = round(float(sua2_val), 2)
                    
                    diferencia = sua1_val - sua2_val
                    diferencia = round(diferencia, 2)
                    resultado[columna] = diferencia
                    
        elif sua1_existe and not sua2_existe:
            pass
        elif not sua1_existe and sua2_existe:
            sua2_data = sua2_row.to_dicts()[0]
            resultado["NOMBRE ASEGURADO"] = sua2_data.get("NOMBRE ASEGURADO", "")
            for col in ["RFC", "CURP", "N_MOVS", "INC", "AUS"]:
                resultado[col] = None
        
        resultados.append(resultado)
    
    result_df = pl.DataFrame(resultados)
    return result_df.to_pandas()

def procesar_hoja_sua_bimestral(sua1_path, sua2_path):
    try:
        sua1_df = pl.read_excel(sua1_path, sheet_name="SUA_BIMESTRAL")
        sua2_df = pl.read_excel(sua2_path, sheet_name="SUA_BIMESTRAL")
    except Exception as e:
        raise ValueError(f"Error al leer las hojas SUA_BIMESTRAL: {e}")
    
    sua1_df = sua1_df.with_columns([
        pl.concat_str([pl.col("RP").cast(pl.Utf8).str.slice(0, 10), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    sua2_df = sua2_df.with_columns([
        pl.concat_str([pl.col("RP").cast(pl.Utf8).str.slice(0, 10), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    todas_ids = set(sua1_df["ID_UNICO"].to_list()) | set(sua2_df["ID_UNICO"].to_list())
    
    resultados = []
    
    for id_unico in todas_ids:
        sua1_row = sua1_df.filter(pl.col("ID_UNICO") == id_unico)
        sua2_row = sua2_df.filter(pl.col("ID_UNICO") == id_unico)
        
        sua1_existe = len(sua1_row) > 0
        sua2_existe = len(sua2_row) > 0
        
        if sua1_existe:
            sua1_data = sua1_row.to_dicts()[0]
            resultado = sua1_data.copy()
        else:
            sua2_data = sua2_row.to_dicts()[0]
            resultado = {col: None for col in sua1_df.columns}
            resultado["RP"] = sua2_data["RP"]
            resultado["NSS"] = sua2_data["NSS"]
            resultado["ID_UNICO"] = sua2_data["ID_UNICO"]
        
        if sua1_existe and sua2_existe:
            sua1_data = sua1_row.to_dicts()[0]
            sua2_data = sua2_row.to_dicts()[0]
            
            resultado["NOMBRE ASEGURADO"] = sua2_data["NOMBRE ASEGURADO"]
            resultado["RFC"] = sua2_data["RFC"]
            resultado["CURP"] = sua2_data["CURP"]
            resultado["N_MOVS"] = sua2_data["N_MOVS"]
            resultado["SDI"] = sua2_data["SDI"]
            resultado["N_CREDITO"] = sua2_data["N_CREDITO"]
            resultado["INC"] = sua1_data.get("INC", 0)
            resultado["AUS"] = sua1_data.get("AUS", 0)
            
            columnas_numericas = ["DIAS", "RETIRO", "CEAV_PAT", "CEAV_OBR", "TOTAL_RCV", 
                                "APORTACION_PAT", "AMORTIZACION", "TOTAL_INF", "TOTAL"]
            
            for columna in columnas_numericas:
                if columna in sua1_data and columna in sua2_data:
                    sua1_val = sua1_data[columna] or 0
                    sua2_val = sua2_data[columna] or 0
                    
                    if isinstance(sua1_val, (float, int)) and isinstance(sua2_val, (float, int)):
                        sua1_val = round(float(sua1_val), 2)
                        sua2_val = round(float(sua2_val), 2)
                    
                    diferencia = sua1_val - sua2_val
                    diferencia = round(diferencia, 2)
                    resultado[columna] = diferencia
                    
        elif sua1_existe and not sua2_existe:
            pass
        elif not sua1_existe and sua2_existe:
            sua2_data = sua2_row.to_dicts()[0]
            resultado["NOMBRE ASEGURADO"] = sua2_data.get("NOMBRE ASEGURADO", "")
            for col in ["RFC", "CURP", "N_MOVS", "N_CREDITO", "INC", "AUS"]:
                resultado[col] = None
        
        resultados.append(resultado)
    
    result_df = pl.DataFrame(resultados)
    return result_df.to_pandas()

def escribir_dataframe_a_excel(worksheet, df, header_fill, header_font):
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
    
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width