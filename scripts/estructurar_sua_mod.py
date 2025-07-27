import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def decode_base62(encoded_str):
    """Decodifica una cadena codificada en base62 a un número entero."""
    base62_chars = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz"
    base = len(base62_chars)
    decoded_value = 0
    for i, char in enumerate(reversed(encoded_str)):
        decoded_value += base62_chars.index(char) * (base ** i)
    return decoded_value


def estructurar_1sua_destino(sua_path, output_folder=None):
    """
    Función modificada para darle forma al archivo .SUA del IMSS, 
    guardando el resultado en la carpeta especificada.
    
    Args:
        sua_path (str): Ruta del archivo .SUA
        output_folder (str): Carpeta donde guardar el resultado. Si es None, usa la carpeta del archivo original.
    
    Returns:
        str: Ruta del archivo Excel creado, o None si hay error
    """
    
    # Creamos una copia y cambiamos el nombre del archivo .SUA a .txt
    if sua_path.endswith('.SUA'):
        new_path = sua_path[:-4] + '.txt'
        with open(sua_path, 'rb') as src, open(new_path, 'wb') as dst:
            dst.write(src.read())
    else:
        raise ValueError("El archivo debe tener la extensión .SUA")
    
    # Obtendremos algunos datos del archivo .txt
    with open(new_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    valor = f"03{content[2:5]}"
    
    # Extraer datos para el nombre personalizado del archivo
    mes = content[30:32]  # Caracteres 31 y 32 (índice 30:32)
    año = content[26:30]  # Caracteres 27 al 30 (índice 26:30)
    registro_patronal = content[2:13].strip()  # Número de registro patronal
    
    # Generar nombre personalizado: "mes-año_registro_patronal"
    nombre_personalizado = f"{mes}-{año}_{registro_patronal}_CEDULA"
    print(f"Nombre personalizado del archivo: {nombre_personalizado}")
    
    # Verificar si el mes es par para crear hoja adicional
    mes_numero = int(mes)
    crear_hoja_adicional = (mes_numero % 2 == 0)
    print(f"Mes: {mes_numero}, ¿Es par?: {crear_hoja_adicional}")

    # Iterar sobre el contenido del archivo .txt para extraer los datos
    registros = []
    try:
        start = 0
        while True:
            start = content.find(valor, start)
            if start == -1:
                break
            
            # Verificar que hay suficientes caracteres para extraer el registro completo
            if start + 295 > len(content):
                break
            
            try:
                # Extraer datos del registro
                rp = content[start+2:start+13].strip()
                nss = content[start+33:start+44].strip()
                rfc = content[start+44:start+57].strip()
                curp = content[start+57:start+75].strip()
                n_credito = content[start+75:start+85].strip()
                if not n_credito:
                    n_credito = None
                movimientos = content[start+93:start+95].strip()
                nombre = content[start+95:start+145].strip()
                dias = int(content[start+154:start+156].strip()) if content[start+154:start+156].strip() else 0
                sdi = float(content[start+145:start+152].strip()) / 100 if content[start+145:start+152].strip() else 0
                incapacidad = int(content[start+156:start+158].strip()) if content[start+156:start+158].strip() else 0
                ausentismo = int(content[start+158:start+160].strip()) if content[start+158:start+160].strip() else 0
                cf = float(content[start+160:start+167].strip()) / 100 if content[start+160:start+167].strip() else 0
                exc_pat = float(content[start+167:start+174].strip()) / 100 if content[start+167:start+174].strip() else 0
                exc_obr = decode_base62(content[start+278:start+281]) / 100 if content[start+278:start+281].strip() else 0
                pd_pat = float(content[start+174:start+181].strip()) / 100 if content[start+174:start+181].strip() else 0
                pd_obr = decode_base62(content[start+281:start+284]) / 100 if content[start+281:start+284].strip() else 0
                gmp_pat = float(content[start+181:start+188].strip()) / 100 if content[start+181:start+188].strip() else 0
                gmp_obr = decode_base62(content[start+284:start+287]) / 100 if content[start+284:start+287].strip() else 0
                rt = float(content[start+188:start+195].strip()) / 100 if content[start+188:start+195].strip() else 0
                iv_pat = float(content[start+195:start+202].strip()) / 100 if content[start+195:start+202].strip() else 0
                iv_obr = decode_base62(content[start+287:start+290]) / 100 if content[start+287:start+290].strip() else 0
                gps = float(content[start+202:start+209].strip()) / 100 if content[start+202:start+209].strip() else 0
                
                total = cf + exc_pat + exc_obr + pd_pat + pd_obr + gmp_pat + gmp_obr + rt + iv_pat + iv_obr + gps
                
                registro = {
                    'RP': rp,
                    'NSS': nss,
                    'NOMBRE ASEGURADO': nombre,
                    'DIAS': dias,
                    'SDI': sdi,
                    'RFC': rfc,
                    'CURP': curp,
                    'N_CREDITO': n_credito,
                    'MOVIMIENTOS': movimientos,
                    'INCAPACIDAD': incapacidad,
                    'AUSENTISMO': ausentismo,
                    'CF': cf,
                    'EXC_PAT': exc_pat,
                    'EXC_OBR': exc_obr,
                    'PD_PAT': pd_pat,
                    'PD_OBR': pd_obr,
                    'GMP_PAT': gmp_pat,
                    'GMP_OBR': gmp_obr,
                    'RT': rt,
                    'IV_PAT': iv_pat,
                    'IV_OBR': iv_obr,
                    'GPS': gps,
                    'TOTAL': total
                }
                
                registros.append(registro)
                
            except (ValueError, IndexError) as e:
                print(f"Error procesando registro en posición {start}: {e}")
                continue
            
            start += 295  # Avanzar al siguiente registro
            
    except Exception as e:
        print(f"Error procesando el archivo: {e}")
        return None
    
    # Procesar registros para SUA_BIMESTRAL (solo si el mes es par)
    registros_suab = []
    if crear_hoja_adicional:
        try:
            start = 0
            while True:
                start = content.find(valor, start)
                if start == -1:
                    break
                
                if start + 295 > len(content):
                    break
                
                try:
                    # Solo agregar campos específicos para SUA_BIMESTRAL
                    rp = content[start+2:start+13].strip()
                    nss = content[start+33:start+44].strip()
                    nombre = content[start+95:start+145].strip()
                    dias = int(content[start+154:start+156].strip()) if content[start+154:start+156].strip() else 0
                    sdi = float(content[start+145:start+152].strip()) / 100 if content[start+145:start+152].strip() else 0
                    n_credito = content[start+75:start+85].strip()
                    if not n_credito:
                        n_credito = None
                    
                    # Campos específicos bimestrales (ejemplo)
                    retiro = float(content[start+209:start+216].strip()) / 100 if content[start+209:start+216].strip() else 0
                    ceav_pat = float(content[start+216:start+223].strip()) / 100 if content[start+216:start+223].strip() else 0
                    ceav_obr = float(content[start+223:start+230].strip()) / 100 if content[start+223:start+230].strip() else 0
                    aportacion_pat = float(content[start+237:start+244].strip()) / 100 if content[start+237:start+244].strip() else 0
                    amortizacion = float(content[start+244:start+251].strip()) / 100 if content[start+244:start+251].strip() else 0
                    
                    registro_suab = {
                        'RP': rp,
                        'NSS': nss,
                        'NOMBRE ASEGURADO': nombre,
                        'DIAS': dias,
                        'SDI': sdi,
                        'N_CREDITO': n_credito,
                        'RETIRO': retiro,
                        'CEAV_PAT': ceav_pat,
                        'CEAV_OBR': ceav_obr,
                        'APORTACION_PAT': aportacion_pat,
                        'AMORTIZACION': amortizacion,
                        'TOTAL': retiro + ceav_pat + ceav_obr + aportacion_pat + amortizacion
                    }

                    registros_suab.append(registro_suab)
                except (ValueError, IndexError) as e:
                    print(f"Error procesando registro SUA_BIMESTRAL en posición {start}: {e}")
                    continue
                
                start += 295  # Avanzar al siguiente registro
        except Exception as e:
            print(f"Error procesando registros SUA_BIMESTRAL: {e}")
            registros_suab = []  # Si hay error, no crear la hoja SUA_BIMESTRAL
    
    # Verificar que se encontraron registros
    if not registros:
        print("No se encontraron registros en el archivo.")
        return None
    
    # Filtrar registros para eliminar aquellos con DIAS = 0
    registros_filtrados = [registro for registro in registros if registro['DIAS'] != 0]
    print(f"Registros antes del filtro: {len(registros)}")
    print(f"Registros después del filtro (eliminando DIAS = 0): {len(registros_filtrados)}")
    
    # Verificar que quedan registros después del filtro
    if not registros_filtrados:
        print("No quedan registros después de filtrar aquellos con DIAS = 0.")
        return None
    
    # Crear un archivo Excel con los datos extraídos
    try:
        df = pd.DataFrame(registros_filtrados)
        
        # Eliminar filas completamente duplicadas
        df = df.drop_duplicates()
        
        # Rellenar valores nulos de N_CREDITO con "-"
        if 'N_CREDITO' in df.columns:
            df['N_CREDITO'] = df['N_CREDITO'].fillna('-').replace('', '-')
        
        # Ordenar por RP y después por NOMBRE ASEGURADO
        df = df.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
        
        # Usar carpeta de destino si se especifica, sino usar carpeta original
        if output_folder:
            base_path = output_folder
        else:
            base_path = os.path.dirname(sua_path)
            
        excel_path = os.path.join(base_path, f"{nombre_personalizado}.xlsx")
        
        # Crear el directorio si no existe
        os.makedirs(base_path, exist_ok=True)
        
        # Guardar el DataFrame en Excel con la hoja SUA_MENSUAL
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='SUA_MENSUAL')
            
            # Si el mes es par, crear la hoja SUA_BIMESTRAL con sus propios registros
            if crear_hoja_adicional and registros_suab:
                registros_suab_filtrados = [registro for registro in registros_suab if registro['DIAS'] != 0]
                print(f"Registros SUA_BIMESTRAL antes del filtro: {len(registros_suab)}")
                print(f"Registros SUA_BIMESTRAL después del filtro (eliminando DIAS = 0): {len(registros_suab_filtrados)}")
                
                if registros_suab_filtrados:
                    df_suab = pd.DataFrame(registros_suab_filtrados)
                    
                    # Eliminar filas completamente duplicadas
                    df_suab = df_suab.drop_duplicates()
                    
                    # Rellenar valores nulos de N_CREDITO con "-"
                    if 'N_CREDITO' in df_suab.columns:
                        df_suab['N_CREDITO'] = df_suab['N_CREDITO'].fillna('-').replace('', '-')
                    
                    # Ordenar por RP y después por NOMBRE ASEGURADO
                    df_suab = df_suab.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
                    
                    df_suab.to_excel(writer, index=False, sheet_name='SUA_BIMESTRAL')
                    print("Hoja SUA_BIMESTRAL creada (mes par detectado)")
                else:
                    print("No hay registros SUA_BIMESTRAL válidos para crear la hoja")
            elif crear_hoja_adicional:
                print("No se pudieron procesar registros SUA_BIMESTRAL")
        
        # Aplicar formato usando openpyxl
        workbook = load_workbook(excel_path)
        
        # Formatear hoja SUA_MENSUAL
        worksheet_suam = workbook['SUA_MENSUAL']
        
        # Definir los estilos para el encabezado SUA_MENSUAL
        header_fill_suam = PatternFill(start_color='611232', end_color='611232', fill_type='solid')
        header_font_suam = Font(color='B3945A', bold=True)
        
        # Aplicar formato a la primera fila de SUA_MENSUAL
        for cell in worksheet_suam[1]:
            cell.fill = header_fill_suam
            cell.font = header_font_suam
        
        # Ajustar el ancho de las columnas en SUA_MENSUAL
        for column in worksheet_suam.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet_suam.column_dimensions[column_letter].width = adjusted_width
        
        # Si existe la hoja SUA_BIMESTRAL, aplicar su formato
        if crear_hoja_adicional and 'SUA_BIMESTRAL' in workbook.sheetnames:
            worksheet_suab = workbook['SUA_BIMESTRAL']
            
            # Definir los estilos para el encabezado SUA_BIMESTRAL
            header_fill_suab = PatternFill(start_color='611232', end_color='611232', fill_type='solid')
            header_font_suab = Font(color='B3945A', bold=True)

            # Aplicar formato a la primera fila de SUA_BIMESTRAL
            for cell in worksheet_suab[1]:
                cell.fill = header_fill_suab
                cell.font = header_font_suab
            
            # Ajustar ancho de columnas para SUA_BIMESTRAL
            for column in worksheet_suab.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet_suab.column_dimensions[column_letter].width = adjusted_width
            
            print("Formato SUA_BIMESTRAL aplicado: fondo #611232, fuente #B3945A")
        
        # Guardar los cambios
        workbook.save(excel_path)
        workbook.close()
        
        print(f"Archivo Excel creado exitosamente: {excel_path}")
        print(f"Se procesaron {len(registros_filtrados)} registros válidos (excluyendo DIAS = 0)")
        print("Formato SUA_MENSUAL aplicado: fondo #611232, fuente #B3945A")
        if crear_hoja_adicional:
            if 'SUA_BIMESTRAL' in workbook.sheetnames:
                print("Hoja SUA_BIMESTRAL incluida (mes par detectado)")
            else:
                print("Mes par detectado pero no se creó hoja SUA_BIMESTRAL")
        
        # Limpiar archivo temporal
        try:
            os.remove(new_path)
        except:
            pass
            
        return excel_path
        
    except Exception as e:
        print(f"Error al crear el archivo Excel: {e}")
        return None


def estructurar_varios_suas(folder_path, output_folder=None):
    """
    Función modificada para procesar múltiples archivos .SUA en una carpeta y combinarlos 
    en un solo archivo Excel con hojas SUA_MENSUAL y SUA_BIMESTRAL (si aplica).
    
    Args:
        folder_path (str): Ruta de la carpeta que contiene los archivos .SUA
        output_folder (str): Carpeta donde guardar el resultado. Si es None, usa folder_path.
    
    Returns:
        str: Ruta del archivo Excel creado, o None si ocurre un error
    """
    
    if not os.path.exists(folder_path):
        print(f"La carpeta {folder_path} no existe.")
        return None
    
    def buscar_archivos_sua(directorio, nivel=0, max_nivel=5):
        """Busca archivos .SUA recursivamente hasta el nivel máximo especificado"""
        archivos_sua = []
        
        if nivel > max_nivel:
            return archivos_sua
        
        try:
            for item in os.listdir(directorio):
                item_path = os.path.join(directorio, item)
                if os.path.isfile(item_path) and item.upper().endswith('.SUA'):
                    archivos_sua.append(item_path)
                elif os.path.isdir(item_path):
                    archivos_sua.extend(buscar_archivos_sua(item_path, nivel + 1, max_nivel))
        except PermissionError:
            print(f"Sin permisos para acceder a: {directorio}")
        except Exception as e:
            print(f"Error accediendo a {directorio}: {e}")
        
        return archivos_sua
    
    # Buscar todos los archivos .SUA en la carpeta y subcarpetas
    print("Buscando archivos .SUA en carpeta y subcarpetas (máximo 5 niveles)...")
    sua_files_paths = buscar_archivos_sua(folder_path)
    
    if not sua_files_paths:
        print("No se encontraron archivos .SUA en la carpeta especificada ni en subcarpetas.")
        return None
    
    print(f"Se encontraron {len(sua_files_paths)} archivos .SUA para procesar:")
    for sua_path in sua_files_paths:
        print(f"  - {sua_path}")
    
    # Listas para almacenar todos los registros
    todos_registros = []
    todos_registros_suab = []
    
    # Variables para determinar el nombre del archivo
    primer_mes = None
    primer_año = None
    crear_hoja_bimestral = False
    
    # Procesar cada archivo .SUA
    for sua_path in sua_files_paths:
        sua_file = os.path.basename(sua_path)
        print(f"Procesando archivo: {sua_file}")
        
        try:
            # Crear archivo .txt temporal
            new_path = sua_path[:-4] + '.txt'
            with open(sua_path, 'rb') as src, open(new_path, 'wb') as dst:
                dst.write(src.read())
            
            # Leer contenido del archivo
            with open(new_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            valor = f"03{content[2:5]}"
            
            # Extraer datos para el nombre del archivo (usar el primer archivo)
            if primer_mes is None:
                primer_mes = content[30:32]
                primer_año = content[26:30]
                crear_hoja_bimestral = (int(primer_mes) % 2 == 0)
            
            # Procesar registros SUA_MENSUAL
            registros_archivo = []
            start = 0
            while True:
                start = content.find(valor, start)
                if start == -1:
                    break
                
                if start + 295 > len(content):
                    break
                
                try:
                    # [Mismo código de extracción que en estructurar_1sua_destino]
                    rp = content[start+2:start+13].strip()
                    nss = content[start+33:start+44].strip()
                    nombre = content[start+95:start+145].strip()
                    dias = int(content[start+154:start+156].strip()) if content[start+154:start+156].strip() else 0
                    
                    # Solo procesar si tiene días
                    if dias > 0:
                        sdi = float(content[start+145:start+152].strip()) / 100 if content[start+145:start+152].strip() else 0
                        rfc = content[start+44:start+57].strip()
                        curp = content[start+57:start+75].strip()
                        n_credito = content[start+75:start+85].strip()
                        if not n_credito:
                            n_credito = None
                        movimientos = content[start+93:start+95].strip()
                        incapacidad = int(content[start+156:start+158].strip()) if content[start+156:start+158].strip() else 0
                        ausentismo = int(content[start+158:start+160].strip()) if content[start+158:start+160].strip() else 0
                        cf = float(content[start+160:start+167].strip()) / 100 if content[start+160:start+167].strip() else 0
                        exc_pat = float(content[start+167:start+174].strip()) / 100 if content[start+167:start+174].strip() else 0
                        exc_obr = decode_base62(content[start+278:start+281]) / 100 if content[start+278:start+281].strip() else 0
                        pd_pat = float(content[start+174:start+181].strip()) / 100 if content[start+174:start+181].strip() else 0
                        pd_obr = decode_base62(content[start+281:start+284]) / 100 if content[start+281:start+284].strip() else 0
                        gmp_pat = float(content[start+181:start+188].strip()) / 100 if content[start+181:start+188].strip() else 0
                        gmp_obr = decode_base62(content[start+284:start+287]) / 100 if content[start+284:start+287].strip() else 0
                        rt = float(content[start+188:start+195].strip()) / 100 if content[start+188:start+195].strip() else 0
                        iv_pat = float(content[start+195:start+202].strip()) / 100 if content[start+195:start+202].strip() else 0
                        iv_obr = decode_base62(content[start+287:start+290]) / 100 if content[start+287:start+290].strip() else 0
                        gps = float(content[start+202:start+209].strip()) / 100 if content[start+202:start+209].strip() else 0
                        
                        total = cf + exc_pat + exc_obr + pd_pat + pd_obr + gmp_pat + gmp_obr + rt + iv_pat + iv_obr + gps
                        
                        registro = {
                            'RP': rp,
                            'NSS': nss,
                            'NOMBRE ASEGURADO': nombre,
                            'DIAS': dias,
                            'SDI': sdi,
                            'RFC': rfc,
                            'CURP': curp,
                            'N_CREDITO': n_credito,
                            'MOVIMIENTOS': movimientos,
                            'INCAPACIDAD': incapacidad,
                            'AUSENTISMO': ausentismo,
                            'CF': cf,
                            'EXC_PAT': exc_pat,
                            'EXC_OBR': exc_obr,
                            'PD_PAT': pd_pat,
                            'PD_OBR': pd_obr,
                            'GMP_PAT': gmp_pat,
                            'GMP_OBR': gmp_obr,
                            'RT': rt,
                            'IV_PAT': iv_pat,
                            'IV_OBR': iv_obr,
                            'GPS': gps,
                            'TOTAL': total
                        }
                        
                        registros_archivo.append(registro)
                        
                except (ValueError, IndexError) as e:
                    print(f"Error procesando registro en {sua_file}, posición {start}: {e}")
                    continue
                
                start += 295
            
            # Agregar registros del archivo actual a la lista total
            todos_registros.extend(registros_archivo)
            
            # Procesar registros SUA_BIMESTRAL si es necesario
            if crear_hoja_bimestral:
                registros_suab_archivo = []
                start = 0
                while True:
                    start = content.find(valor, start)
                    if start == -1:
                        break
                    
                    if start + 295 > len(content):
                        break
                    
                    try:
                        dias = int(content[start+154:start+156].strip()) if content[start+154:start+156].strip() else 0
                        
                        if dias > 0:
                            rp = content[start+2:start+13].strip()
                            nss = content[start+33:start+44].strip()
                            nombre = content[start+95:start+145].strip()
                            sdi = float(content[start+145:start+152].strip()) / 100 if content[start+145:start+152].strip() else 0
                            n_credito = content[start+75:start+85].strip()
                            if not n_credito:
                                n_credito = None
                            
                            retiro = float(content[start+209:start+216].strip()) / 100 if content[start+209:start+216].strip() else 0
                            ceav_pat = float(content[start+216:start+223].strip()) / 100 if content[start+216:start+223].strip() else 0
                            ceav_obr = float(content[start+223:start+230].strip()) / 100 if content[start+223:start+230].strip() else 0
                            aportacion_pat = float(content[start+237:start+244].strip()) / 100 if content[start+237:start+244].strip() else 0
                            amortizacion = float(content[start+244:start+251].strip()) / 100 if content[start+244:start+251].strip() else 0
                            
                            registro_suab = {
                                'RP': rp,
                                'NSS': nss,
                                'NOMBRE ASEGURADO': nombre,
                                'DIAS': dias,
                                'SDI': sdi,
                                'N_CREDITO': n_credito,
                                'RETIRO': retiro,
                                'CEAV_PAT': ceav_pat,
                                'CEAV_OBR': ceav_obr,
                                'APORTACION_PAT': aportacion_pat,
                                'AMORTIZACION': amortizacion,
                                'TOTAL': retiro + ceav_pat + ceav_obr + aportacion_pat + amortizacion
                            }
                            
                            registros_suab_archivo.append(registro_suab)
                            
                    except (ValueError, IndexError) as e:
                        print(f"Error procesando registro SUA_BIMESTRAL en {sua_file}, posición {start}: {e}")
                        continue
                    
                    start += 295
                
                todos_registros_suab.extend(registros_suab_archivo)
                print(f"Procesado {sua_file}: {len(registros_suab_archivo)} registros SUA_BIMESTRAL")
            
            print(f"Procesado {sua_file}: {len(registros_archivo)} registros SUA_MENSUAL")
            
            # Limpiar archivo temporal
            try:
                os.remove(new_path)
            except:
                pass
                
        except Exception as e:
            print(f"Error procesando archivo {sua_file}: {e}")
            continue
    
    # Verificar que se encontraron registros
    if not todos_registros:
        print("No se encontraron registros válidos en ningún archivo.")
        return None
    
    # Filtrar registros para eliminar aquellos con DIAS = 0
    registros_filtrados = [registro for registro in todos_registros if registro['DIAS'] != 0]
    print(f"Total registros antes del filtro: {len(todos_registros)}")
    print(f"Total registros después del filtro (eliminando DIAS = 0): {len(registros_filtrados)}")
    
    if not registros_filtrados:
        print("No quedan registros después de filtrar aquellos con DIAS = 0.")
        return None
    
    # Usar carpeta de destino si se especifica, sino usar folder_path
    if output_folder:
        base_path = output_folder
    else:
        base_path = folder_path
    
    # Generar nombre del archivo
    nombre_personalizado = f"{primer_mes}-{primer_año}_MULTI_CEDULA"
    excel_path = os.path.join(base_path, f"{nombre_personalizado}.xlsx")
    
    # Crear el directorio si no existe
    os.makedirs(base_path, exist_ok=True)
    
    try:
        # Crear DataFrame principal
        df = pd.DataFrame(registros_filtrados)
        
        # Eliminar filas completamente duplicadas
        df = df.drop_duplicates()
        
        # Rellenar valores nulos de N_CREDITO con "-"
        if 'N_CREDITO' in df.columns:
            df['N_CREDITO'] = df['N_CREDITO'].fillna('-').replace('', '-')
        
        # Ordenar por RP y después por NOMBRE ASEGURADO
        df = df.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
        
        # Guardar el archivo Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='SUA_MENSUAL')
            
            # Crear hoja SUA_BIMESTRAL si es necesario
            if crear_hoja_bimestral and todos_registros_suab:
                registros_suab_filtrados = [registro for registro in todos_registros_suab if registro['DIAS'] != 0]
                print(f"Total registros SUA_BIMESTRAL antes del filtro: {len(todos_registros_suab)}")
                print(f"Total registros SUA_BIMESTRAL después del filtro: {len(registros_suab_filtrados)}")
                
                if registros_suab_filtrados:
                    df_suab = pd.DataFrame(registros_suab_filtrados)
                    
                    # Eliminar filas completamente duplicadas
                    df_suab = df_suab.drop_duplicates()
                    
                    # Rellenar valores nulos de N_CREDITO con "-"
                    if 'N_CREDITO' in df_suab.columns:
                        df_suab['N_CREDITO'] = df_suab['N_CREDITO'].fillna('-').replace('', '-')
                    
                    # Ordenar por RP y después por NOMBRE ASEGURADO
                    df_suab = df_suab.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
                    df_suab.to_excel(writer, index=False, sheet_name='SUA_BIMESTRAL')
                    print("Hoja SUA_BIMESTRAL creada")
                else:
                    print("No hay registros SUA_BIMESTRAL válidos para crear la hoja")
            elif crear_hoja_bimestral:
                print("No se pudieron procesar registros SUA_BIMESTRAL")
        
        # Aplicar formato usando openpyxl
        workbook = load_workbook(excel_path)
        
        # Formatear hoja SUA_MENSUAL
        worksheet_suam = workbook['SUA_MENSUAL']
        header_fill_suam = PatternFill(start_color='611232', end_color='611232', fill_type='solid')
        header_font_suam = Font(color='B3945A', bold=True)
        
        for cell in worksheet_suam[1]:
            cell.fill = header_fill_suam
            cell.font = header_font_suam
        
        # Ajustar ancho de columnas SUA_MENSUAL
        for column in worksheet_suam.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet_suam.column_dimensions[column_letter].width = adjusted_width
        
        # Formatear hoja SUA_BIMESTRAL si existe
        if crear_hoja_bimestral and 'SUA_BIMESTRAL' in workbook.sheetnames:
            worksheet_suab = workbook['SUA_BIMESTRAL']
            header_fill_suab = PatternFill(start_color='611232', end_color='611232', fill_type='solid')
            header_font_suab = Font(color='B3945A', bold=True)
            
            for cell in worksheet_suab[1]:
                cell.fill = header_fill_suab
                cell.font = header_font_suab
            
            # Ajustar ancho de columnas para SUA_BIMESTRAL
            for column in worksheet_suab.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet_suab.column_dimensions[column_letter].width = adjusted_width
            
            print("Formato SUA_BIMESTRAL aplicado: fondo #611232, fuente #B3945A")
        
        # Guardar cambios
        workbook.save(excel_path)
        workbook.close()
        
        print(f"\n=== RESUMEN ===")
        print(f"Archivo Excel creado exitosamente: {excel_path}")
        print(f"Archivos .SUA procesados: {len(sua_files_paths)}")
        print(f"Total registros válidos procesados: {len(registros_filtrados)}")
        print("Formato aplicado: fondo #611232, fuente #B3945A")
        if crear_hoja_bimestral:
            if 'SUA_BIMESTRAL' in workbook.sheetnames:
                print("Hoja SUA_BIMESTRAL incluida (mes par detectado)")
            else:
                print("Mes par detectado pero no se creó hoja SUA_BIMESTRAL")
        
        return excel_path
        
    except Exception as e:
        print(f"Error al crear el archivo Excel: {e}")
        return None
