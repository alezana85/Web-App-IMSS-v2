import polars as pl
import pandas as pd
import os
import re
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook

def sua_vs_emision(sua_path, emision_path, archive_path):
    """
    Compara archivos SUA vs Emisión y genera un archivo Excel con las diferencias.
    
    Args:
        sua_path (str): Ruta del archivo SUA
        emision_path (str): Ruta del archivo de emisión
        archive_path (str): Ruta donde guardar el archivo resultado
    """
    
    # Extraer período de los nombres de archivo
    sua_filename = os.path.basename(sua_path)
    emision_filename = os.path.basename(emision_path)
    
    # Extraer mes y año del nombre del archivo SUA
    sua_match = re.search(r'(\d{2})-(\d{4})', sua_filename)
    emision_match = re.search(r'(\d{2})-(\d{4})', emision_filename)
    
    if not sua_match or not emision_match:
        raise ValueError("Los nombres de archivo deben contener el formato MM-YYYY")
    
    sua_mes, sua_año = sua_match.groups()
    emision_mes, emision_año = emision_match.groups()
    
    # Verificar que ambos archivos sean del mismo período
    if sua_mes != emision_mes or sua_año != emision_año:
        raise ValueError(f"Los archivos no corresponden al mismo período. SUA: {sua_mes}-{sua_año}, Emisión: {emision_mes}-{emision_año}")
    
    mes = int(sua_mes)
    año = int(sua_año)
    
    # Determinar número de hojas según el mes (par = 2 hojas, impar = 1 hoja)
    num_hojas = 2 if mes % 2 == 0 else 1
    
    # Crear archivo Excel resultado
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)  # Remover hoja por defecto
    
    # Colores para formato (morado y verde del Joker)
    purple_fill = PatternFill(start_color="8B008B", end_color="8B008B", fill_type="solid")
    green_font = Font(color="00FF00", bold=True)
    
    # Procesar hoja 1 (MENSUAL)
    hoja1_result = procesar_hoja_mensual(sua_path, emision_path)
    ws1 = wb.create_sheet("MENSUAL")
    escribir_dataframe_a_excel(ws1, hoja1_result, purple_fill, green_font)
    
    # Procesar hoja 2 (BIMESTRAL) si es necesario
    if num_hojas == 2:
        hoja2_result = procesar_hoja_bimestral(sua_path, emision_path)
        ws2 = wb.create_sheet("BIMESTRAL")
        escribir_dataframe_a_excel(ws2, hoja2_result, purple_fill, green_font)
    
    # Guardar archivo
    output_filename = f"{sua_mes}_{sua_año}_CONFRONTA.xlsx"
    output_path = os.path.join(archive_path, output_filename)
    wb.save(output_path)
    
    return output_path


def procesar_hoja_mensual(sua_path, emision_path):
    """Procesa la comparación de la hoja mensual (hoja 1)"""
    
    # Leer archivos con polars
    try:
        sua_df = pl.read_excel(sua_path, sheet_id=1)
        emision_df = pl.read_excel(emision_path, sheet_id=1)
    except Exception as e:
        raise ValueError(f"Error al leer los archivos Excel: {e}")
    
    # Crear columna compuesta para identificación única
    sua_df = sua_df.with_columns([
        pl.concat_str([pl.col("RP"), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    emision_df = emision_df.with_columns([
        pl.concat_str([pl.col("RP"), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    # Obtener todas las IDs únicas de ambos archivos
    todas_ids = set(sua_df["ID_UNICO"].to_list()) | set(emision_df["ID_UNICO"].to_list())
    
    resultados = []
    
    for id_unico in todas_ids:
        sua_row = sua_df.filter(pl.col("ID_UNICO") == id_unico)
        emision_row = emision_df.filter(pl.col("ID_UNICO") == id_unico)
        
        sua_existe = len(sua_row) > 0
        emision_existe = len(emision_row) > 0
        
        # Preparar fila resultado
        if sua_existe:
            sua_data = sua_row.to_dicts()[0]
            resultado = sua_data.copy()
        else:
            emision_data = emision_row.to_dicts()[0]
            resultado = {col: None for col in sua_df.columns}
            resultado["RP"] = emision_data["RP"]
            resultado["NSS"] = emision_data["NSS"]
            resultado["ID_UNICO"] = emision_data["ID_UNICO"]
        
        observaciones = []
        
        # Procesar según disponibilidad de datos
        if sua_existe and emision_existe:
            sua_data = sua_row.to_dicts()[0]
            emision_data = emision_row.to_dicts()[0]
            
            # Comparar NOMBRE_ASEGURADO
            if "NOMBRE_ASEGURADO" in sua_data and "NOMBRE_ASEGURADO" in emision_data:
                if sua_data["NOMBRE_ASEGURADO"] != emision_data["NOMBRE_ASEGURADO"]:
                    observaciones.append("NOMBRE DIFERENTE")
                resultado["NOMBRE_ASEGURADO"] = emision_data["NOMBRE_ASEGURADO"]
            
            # Comparar y calcular diferencias para columnas numéricas
            columnas_comparacion = ["DIAS", "SDI", "CF", "EXC_PAT", "EXC_OBR", "PD_PAT", "PD_OBR", 
                                  "GMP_PAT", "GMP_OBR", "RT", "IV_PAT", "IV_OBR", "GPS", "TOTAL"]
            
            diferencias_cuotas = False
            diferencia_dias = False
            diferencia_sdi = False
            diferencia_total = 0
            
            for columna in columnas_comparacion:
                if columna in sua_data and columna in emision_data:
                    sua_val = sua_data[columna] or 0
                    emision_val = emision_data[columna] or 0
                    diferencia = sua_val - emision_val
                    
                    # Aplicar tolerancia para errores de precisión de punto flotante
                    if abs(diferencia) < 1e-10:
                        diferencia = 0
                    
                    if columna == "DIAS":
                        resultado[columna] = diferencia
                        if diferencia != 0:
                            diferencia_dias = True
                            if diferencia > 0:
                                observaciones.append("MAS DIAS EN SUA")
                            else:
                                observaciones.append("MAS DIAS EN EMISION")
                    elif columna == "SDI":
                        resultado[columna] = diferencia
                        if diferencia != 0:
                            diferencia_sdi = True
                            observaciones.append("SALARIO DIFERENTE")
                    elif columna == "TOTAL":
                        resultado[columna] = diferencia
                        diferencia_total = diferencia
                    else:
                        resultado[columna] = diferencia
                        if diferencia != 0:
                            diferencias_cuotas = True
            
            # Verificar diferencias por incapacidad/ausentismo
            if diferencia_total != 0:
                inc_val = sua_data.get("INC", 0) or 0
                aus_val = sua_data.get("AUS", 0) or 0
                
                if inc_val > 0 or aus_val > 0:
                    dias_emision = emision_data.get("DIAS", 0) or 0
                    total_emision = emision_data.get("TOTAL", 0) or 0
                    
                    if dias_emision > 0:
                        calculo_incapacidad = (total_emision / dias_emision) * (inc_val + aus_val)
                        if abs(abs(diferencia_total) - abs(calculo_incapacidad)) < 0.40:  # Tolerancia para cálculo de incapacidad
                            observaciones = [obs for obs in observaciones if "DIFERENCIAS" not in obs]
                            observaciones.append("SIN DIFERENCIAS POR INCAPACIDAD/AUSENTISMO")
                        else:
                            if not any("DIFERENCIAS" in obs for obs in observaciones):
                                observaciones.append("DIFERENCIAS")
                    else:
                        if not any("DIFERENCIAS" in obs for obs in observaciones):
                            observaciones.append("DIFERENCIAS")
                else:
                    if not any("DIFERENCIAS" in obs for obs in observaciones):
                        observaciones.append("DIFERENCIAS")
            
            # Mantener valores de INC y AUS de SUA
            resultado["INC"] = sua_data.get("INC", 0)
            resultado["AUS"] = sua_data.get("AUS", 0)
            
        elif sua_existe and not emision_existe:
            observaciones.append("NO APARECE EN EMISION")
        elif not sua_existe and emision_existe:
            emision_data = emision_row.to_dicts()[0]
            observaciones.append("NO APARECE EN SUA")
            resultado["NOMBRE_ASEGURADO"] = emision_data.get("NOMBRE_ASEGURADO", "")
            # Valores nulos para columnas que no existen en emisión
            for col in ["RFC", "CURP", "N_MOVS", "INC", "AUS"]:
                resultado[col] = None
        
        # Establecer observaciones finales
        if not observaciones:
            resultado["OBSERVACIONES"] = "SIN DIFERENCIAS"
        else:
            resultado["OBSERVACIONES"] = ", ".join(observaciones)
        
        resultados.append(resultado)
    
    # Convertir a DataFrame de polars
    if resultados:
        # Remover ID_UNICO antes de retornar
        for resultado in resultados:
            resultado.pop("ID_UNICO", None)
        
        resultado_df = pl.DataFrame(resultados)
        
        # Ordenar por RP y después por NOMBRE ASEGURADO
        resultado_df = resultado_df.sort(["RP", "NOMBRE ASEGURADO"])
        
        return resultado_df
    else:
        return pl.DataFrame()


def procesar_hoja_bimestral(sua_path, emision_path):
    """Procesa la comparación de la hoja bimestral (hoja 2)"""
    
    # Leer archivos con polars
    try:
        sua_df = pl.read_excel(sua_path, sheet_id=2)
        emision_df = pl.read_excel(emision_path, sheet_id=2)
    except Exception as e:
        raise ValueError(f"Error al leer las hojas bimestrales: {e}")
    
    # Crear columna compuesta para identificación única
    sua_df = sua_df.with_columns([
        pl.concat_str([pl.col("RP"), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    emision_df = emision_df.with_columns([
        pl.concat_str([pl.col("RP"), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    # Obtener todas las IDs únicas de ambos archivos
    todas_ids = set(sua_df["ID_UNICO"].to_list()) | set(emision_df["ID_UNICO"].to_list())
    
    resultados = []
    
    for id_unico in todas_ids:
        sua_row = sua_df.filter(pl.col("ID_UNICO") == id_unico)
        emision_row = emision_df.filter(pl.col("ID_UNICO") == id_unico)
        
        sua_existe = len(sua_row) > 0
        emision_existe = len(emision_row) > 0
        
        # Preparar fila resultado
        if sua_existe:
            sua_data = sua_row.to_dicts()[0]
            resultado = sua_data.copy()
        else:
            emision_data = emision_row.to_dicts()[0]
            resultado = {col: None for col in sua_df.columns}
            resultado["RP"] = emision_data["RP"]
            resultado["NSS"] = emision_data["NSS"]
            resultado["ID_UNICO"] = emision_data["ID_UNICO"]
        
        observaciones = []
        
        # Procesar según disponibilidad de datos
        if sua_existe and emision_existe:
            sua_data = sua_row.to_dicts()[0]
            emision_data = emision_row.to_dicts()[0]
            
            # Comparar NOMBRE_ASEGURADO
            if "NOMBRE_ASEGURADO" in sua_data and "NOMBRE_ASEGURADO" in emision_data:
                if sua_data["NOMBRE_ASEGURADO"] != emision_data["NOMBRE_ASEGURADO"]:
                    observaciones.append("NOMBRE DIFERENTE")
                resultado["NOMBRE_ASEGURADO"] = emision_data["NOMBRE_ASEGURADO"]
            
            # Comparar N_CREDITO
            if "N_CREDITO" in sua_data and "N_CREDITO" in emision_data:
                sua_credito = sua_data["N_CREDITO"] or "-"
                emision_credito = emision_data["N_CREDITO"] or "-"
                
                # Solo comparar si ambos tienen valores diferentes a "-"
                if sua_credito != "-" and emision_credito != "-" and sua_credito != emision_credito:
                    observaciones.append("NUMERO DE CREDITO DIFERENTE")
                
                # Mantener el valor de emisión si existe, sino el de SUA
                resultado["N_CREDITO"] = emision_credito if emision_credito != "-" else sua_credito
            
            # Comparar y calcular diferencias para columnas numéricas
            columnas_comparacion = ["DIAS", "SDI", "RETIRO", "CEAV_PAT", "CEAV_OBR", "TOTAL_RCV", 
                                  "APORTACION_PAT", "AMORTIZACION", "TOTAL_INF", "TOTAL"]
            
            diferencias_cuotas = False
            diferencia_dias = False
            diferencia_sdi = False
            diferencia_total = 0
            diferencia_total_rcv = 0
            
            for columna in columnas_comparacion:
                if columna in sua_data and columna in emision_data:
                    sua_val = sua_data[columna] or 0
                    emision_val = emision_data[columna] or 0
                    diferencia = sua_val - emision_val
                    
                    # Aplicar tolerancia para errores de precisión de punto flotante
                    if abs(diferencia) < 1e-10:
                        diferencia = 0
                    
                    if columna == "DIAS":
                        resultado[columna] = diferencia
                        if diferencia != 0:
                            diferencia_dias = True
                            if diferencia > 0:
                                observaciones.append("MAS DIAS EN SUA")
                            else:
                                observaciones.append("MAS DIAS EN EMISION")
                    elif columna == "SDI":
                        resultado[columna] = diferencia
                        if diferencia != 0:
                            diferencia_sdi = True
                            observaciones.append("SALARIO DIFERENTE")
                    elif columna == "TOTAL":
                        resultado[columna] = diferencia
                        diferencia_total = diferencia
                    elif columna == "TOTAL_RCV":
                        resultado[columna] = diferencia
                        diferencia_total_rcv = diferencia
                    else:
                        resultado[columna] = diferencia
                        if diferencia != 0:
                            diferencias_cuotas = True
            
            # Verificar diferencias por incapacidad/ausentismo en bimestral
            if diferencia_total != 0:
                inc_val = sua_data.get("INC", 0) or 0
                aus_val = sua_data.get("AUS", 0) or 0
                
                if inc_val > 0 or aus_val > 0:
                    dias_emision = emision_data.get("DIAS", 0) or 0
                    ceav_pat_emision = emision_data.get("CEAV_PAT", 0) or 0
                    ceav_obr_emision = emision_data.get("CEAV_OBR", 0) or 0
                    
                    # Verificar si hay diferencia en AMORTIZACION
                    diferencia_amortizacion = 0
                    if "AMORTIZACION" in sua_data and "AMORTIZACION" in emision_data:
                        amortizacion_sua = sua_data.get("AMORTIZACION", 0) or 0
                        amortizacion_emision = emision_data.get("AMORTIZACION", 0) or 0
                        diferencia_amortizacion = amortizacion_sua - amortizacion_emision
                        if abs(diferencia_amortizacion) < 1e-10:
                            diferencia_amortizacion = 0
                    
                    if dias_emision > 0:
                        calculo_incapacidad = ((ceav_pat_emision + ceav_obr_emision) / dias_emision) * (inc_val + aus_val)
                        
                        # Solo aplicar "SIN DIFERENCIAS POR INCAPACIDAD/AUSENTISMO" si no hay diferencia en AMORTIZACION
                        if (abs(abs(diferencia_total_rcv) - abs(calculo_incapacidad)) < 0.40 and 
                            diferencia_amortizacion == 0):
                            observaciones = [obs for obs in observaciones if "DIFERENCIAS" not in obs]
                            observaciones.append("SIN DIFERENCIAS POR INCAPACIDAD/AUSENTISMO")
                        else:
                            if not any("DIFERENCIAS" in obs for obs in observaciones):
                                observaciones.append("DIFERENCIAS")
                    else:
                        if not any("DIFERENCIAS" in obs for obs in observaciones):
                            observaciones.append("DIFERENCIAS")
                else:
                    if not any("DIFERENCIAS" in obs for obs in observaciones):
                        observaciones.append("DIFERENCIAS")
            
            # Mantener valores de INC y AUS de SUA
            resultado["INC"] = sua_data.get("INC", 0)
            resultado["AUS"] = sua_data.get("AUS", 0)
            
            # Mantener valores que solo existen en emisión
            resultado["T_CREDITO"] = emision_data.get("T_CREDITO", "")
            resultado["V_CREDITO"] = emision_data.get("V_CREDITO", 0)
            
        elif sua_existe and not emision_existe:
            observaciones.append("NO APARECE EN EMISION")
        elif not sua_existe and emision_existe:
            emision_data = emision_row.to_dicts()[0]
            observaciones.append("NO APARECE EN SUA")
            resultado["NOMBRE_ASEGURADO"] = emision_data.get("NOMBRE_ASEGURADO", "")
            resultado["T_CREDITO"] = emision_data.get("T_CREDITO", "")
            resultado["V_CREDITO"] = emision_data.get("V_CREDITO", 0)
            resultado["N_CREDITO"] = emision_data.get("N_CREDITO", "")
            # Valores nulos para columnas que no existen en emisión
            for col in ["RFC", "CURP", "N_MOVS", "INC", "AUS"]:
                resultado[col] = None
        
        # Establecer observaciones finales
        if not observaciones:
            resultado["OBSERVACIONES"] = "SIN DIFERENCIAS"
        else:
            resultado["OBSERVACIONES"] = ", ".join(observaciones)
        
        resultados.append(resultado)
    
    # Convertir a DataFrame de polars
    if resultados:
        # Remover ID_UNICO antes de retornar
        for resultado in resultados:
            resultado.pop("ID_UNICO", None)
        
        resultado_df = pl.DataFrame(resultados)
        
        # Ordenar por RP y después por NOMBRE ASEGURADO
        resultado_df = resultado_df.sort(["RP", "NOMBRE ASEGURADO"])
        
        # Reordenar columnas para que T_CREDITO y V_CREDITO vayan antes de N_CREDITO
        columnas_originales = resultado_df.columns
        columnas_reordenadas = []
        
        for col in columnas_originales:
            if col not in ["T_CREDITO", "V_CREDITO", "N_CREDITO"]:
                columnas_reordenadas.append(col)
            elif col == "N_CREDITO":
                # Insertar T_CREDITO y V_CREDITO antes de N_CREDITO
                if "T_CREDITO" in columnas_originales:
                    columnas_reordenadas.append("T_CREDITO")
                if "V_CREDITO" in columnas_originales:
                    columnas_reordenadas.append("V_CREDITO")
                columnas_reordenadas.append("N_CREDITO")
        
        # Agregar cualquier columna faltante al final
        for col in ["T_CREDITO", "V_CREDITO"]:
            if col in columnas_originales and col not in columnas_reordenadas:
                columnas_reordenadas.append(col)
        
        resultado_df = resultado_df.select(columnas_reordenadas)
        
        return resultado_df
    else:
        return pl.DataFrame()


def escribir_dataframe_a_excel(worksheet, dataframe, purple_fill, green_font):
    """Escribe un DataFrame de polars a una hoja de Excel con formato"""
    
    if dataframe.is_empty():
        return
    
    # Convertir a pandas para facilitar escritura a Excel
    df_pandas = dataframe.to_pandas()
    
    # Escribir encabezados
    for col_idx, column_name in enumerate(df_pandas.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = column_name
        cell.fill = purple_fill
        cell.font = green_font
    
    # Escribir datos
    for row_idx, row in enumerate(df_pandas.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            worksheet.cell(row=row_idx, column=col_idx).value = value