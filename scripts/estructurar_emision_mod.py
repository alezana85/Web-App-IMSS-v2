import os
import pandas as pd
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


def estrucurar_varias_emisiones_destino(folder_path, output_folder=None):
    """
    Función modificada para procesar múltiples archivos de emisión en una carpeta y combinarlos 
    en un solo archivo Excel con hojas EMA y EBA (si aplica), guardando en carpeta específica.
    
    Args:
        folder_path (str): Ruta de la carpeta que contiene los archivos de emisión
        output_folder (str): Carpeta donde guardar el resultado. Si es None, usa folder_path.
    
    Returns:
        str: Ruta del archivo Excel creado, o None si ocurre un error
    """
    
    if not os.path.exists(folder_path):
        print(f"La carpeta {folder_path} no existe.")
        return None
    
    def buscar_archivos_emision(directorio, nivel=0, max_nivel=5):
        """Busca archivos .xls de emisión recursivamente hasta el nivel máximo especificado"""
        archivos_emision = []
        
        if nivel > max_nivel:
            return archivos_emision
        
        try:
            for item in os.listdir(directorio):
                item_path = os.path.join(directorio, item)
                if os.path.isfile(item_path) and item.lower().endswith('.xls'):
                    # Verificar que es un archivo de emisión válido
                    try:
                        wb = xlrd.open_workbook(item_path)
                        if wb.nsheets >= 2:  # Debe tener al menos 2 hojas
                            archivos_emision.append(item_path)
                    except:
                        pass  # Ignorar archivos que no se pueden abrir
                elif os.path.isdir(item_path):
                    archivos_emision.extend(buscar_archivos_emision(item_path, nivel + 1, max_nivel))
        except PermissionError:
            print(f"Sin permisos para acceder a: {directorio}")
        except Exception as e:
            print(f"Error al buscar en {directorio}: {e}")
        
        return archivos_emision
    
    # Buscar todos los archivos de emisión en la carpeta y subcarpetas
    print("Buscando archivos de emisión (.xls) en carpeta y subcarpetas (máximo 5 niveles)...")
    archivos_emision_paths = buscar_archivos_emision(folder_path)
    
    if not archivos_emision_paths:
        print("No se encontraron archivos de emisión válidos en la carpeta especificada ni en subcarpetas.")
        return None
    
    print(f"Se encontraron {len(archivos_emision_paths)} archivos de emisión para procesar:")
    for archivo_path in archivos_emision_paths:
        print(f"  - {archivo_path}")
    
    # Listas para almacenar todos los registros
    todos_ema = []
    todos_eba = []
    
    # Variables para determinar el nombre del archivo
    primer_mes = None
    primer_anio = None
    crear_hoja_eba = False
    
    # Procesar cada archivo de emisión
    for archivo_path in archivos_emision_paths:
        archivo_nombre = os.path.basename(archivo_path)
        print(f"Procesando archivo: {archivo_nombre}")
        
        try:
            # Abrir archivo para obtener información del período
            wb = xlrd.open_workbook(archivo_path)
            ws = wb.sheet_by_index(0)
            
            # Verificar el periodo en la celda B8
            periodo = ws.cell_value(7, 1)  # B8 = fila 7, columna 1
            if isinstance(periodo, str):
                mes, anio = periodo.split('/')
                mes = int(mes)
                anio = int(anio)
            else:
                continue
            
            # Guardar primer mes/año encontrado para el nombre del archivo
            if primer_mes is None:
                primer_mes = mes
                primer_anio = anio
            
            # Obtener el valor de RP de la celda B9
            rp_value = ws.cell_value(8, 1)
            
            # Leer la hoja 2 (Emision Mensual)
            ema = pd.read_excel(archivo_path, sheet_name=1, header=4, engine='xlrd')
            
            # Procesar EMA
            ema = ema[ema['Tipo del Movimiento'] != 2]
            ema['Nombre'] = ema['Nombre'].str.replace('#', 'Ñ', regex=False).str.rstrip()
            ema = ema.drop(columns=['Origen del Movimiento', 'Tipo del Movimiento'], errors='ignore')
            ema['NSS'] = ema['NSS'].apply(lambda x: str(int(x)).zfill(11) if pd.notna(x) else x)
            ema['Fecha del Movimiento'] = ema['Fecha del Movimiento'].replace('-', pd.NaT)
            ema['Fecha del Movimiento'] = pd.to_datetime(ema['Fecha del Movimiento'], errors='coerce')
            ema = ema.sort_values(['NSS', 'Fecha del Movimiento'], na_position='last')

            ema = ema.groupby('NSS').agg({
                'Nombre': 'first',
                'Días': 'sum',
                'Salario Diario': 'last',
                'Cuota Fija': 'sum',
                'Excedente Patronal': 'sum',
                'Excedente Obrero': 'sum',
                'Prestaciones en Dinero Patronal': 'sum',
                'Prestaciones en Dinero Obrero': 'sum',
                'Gastos Médicos y Pensionados Patronal': 'sum',
                'Gastos Médicos y Pensionados Obrero': 'sum',
                'Riesgos de Trabajo': 'sum',
                'Invalidez y Vida Patronal': 'sum',
                'Invalidez y Vida Obrero': 'sum',
                'Guarderías y Prestaciones Sociales': 'sum',
                'Total': 'sum'
            }).reset_index()

            # Agregar la columna RP como primera columna
            ema.insert(0, 'RP', rp_value)

            # Cambiar nombre a las columnas
            ema.columns = [
                'RP', 'NSS', 'NOMBRE ASEGURADO', 'DIAS', 'SDI', 'CF',
                'EXC_PAT', 'EXC_OBR', 'PD_PAT', 'PD_OBR', 'GMP_PAT',
                'GMP_OBR', 'RT', 'IV_PAT', 'IV_OBR', 'GPS', 'TOTAL'
            ]

            # Ordenar el DataFrame por RP y después por NOMBRE ASEGURADO
            ema = ema.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
            todos_ema.append(ema)
            
            # Si el mes es par, procesar también la hoja 3
            if mes % 2 == 0:
                crear_hoja_eba = True
                print(f"Mes par ({mes}) detectado, procesando también la hoja 3...")
                
                # Leer la hoja 3
                eba = pd.read_excel(archivo_path, sheet_name=2, header=4, engine='xlrd')
                
                # Procesar EBA
                eba = eba[eba['Tipo del Movimiento'] != 2]
                eba['Nombre'] = eba['Nombre'].str.replace('#', 'Ñ', regex=False).str.rstrip()
                eba = eba.drop(columns=['Origen del Movimiento', 'Tipo del Movimiento'], errors='ignore')
                eba['NSS'] = eba['NSS'].apply(lambda x: str(int(x)).zfill(11) if pd.notna(x) else x)
                eba['Número de Crédito'] = eba['Número de Crédito'].apply(
                    lambda x: str(x)[-10:] if pd.notna(x) else x
                )
                eba['Fecha del Movimiento'] = eba['Fecha del Movimiento'].replace('-', pd.NaT)
                eba['Fecha del Movimiento'] = pd.to_datetime(eba['Fecha del Movimiento'], errors='coerce')
                eba = eba.sort_values(['NSS', 'Fecha del Movimiento'], na_position='last')

                eba = eba.groupby('NSS').agg({
                    'Nombre': 'first',
                    'Días': 'sum',
                    'Salario Diario': 'last',
                    'Retiro': 'sum',
                    'Cesantía en Edad Avanzada y Vejez Patronal': 'sum',
                    'Cesantía en Edad Avanzada y Vejez Obrero': 'sum',
                    'Subtotal RCV': 'sum',
                    'Aportación Patronal': 'sum',
                    'Tipo de Descuento': 'last',
                    'Valor de Descuento': 'last',
                    'Número de Crédito': 'last',
                    'Amortización': 'sum',
                    'Subtotal Infonavit': 'sum',
                    'Total': 'sum'
                }).reset_index()

                # Agregar la columna RP como primera columna
                eba.insert(0, 'RP', rp_value)

                # Cambiar nombre a las columnas
                eba.columns = [
                    'RP', 'NSS', 'NOMBRE ASEGURADO', 'DIAS', 'SDI', 'RETIRO',
                    'CEAV_PAT', 'CEAV_OBR', 'TOTAL_RCV', 'APORTACION_PAT', 'T_CREDITO',
                    'V_CREDITO', 'N_CREDITO', 'AMORTIZACION', 'TOTAL_INF', 'TOTAL'
                ]

                # Ordenar el DataFrame por RP y después por NOMBRE ASEGURADO
                eba = eba.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
                todos_eba.append(eba)
                
        except Exception as e:
            print(f"Error procesando {archivo_nombre}: {e}")
            continue
    
    # Verificar que se encontraron registros
    if not todos_ema:
        print("No se encontraron registros EMA válidos en ningún archivo.")
        return None
    
    # Combinar todos los DataFrames EMA
    ema_combinado = pd.concat(todos_ema, ignore_index=True)
    
    # Ordenar por RP y después por NOMBRE ASEGURADO
    ema_combinado = ema_combinado.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
    
    print(f"Total registros EMA combinados: {len(ema_combinado)}")
    
    # Combinar todos los DataFrames EBA si existen
    eba_combinado = None
    if crear_hoja_eba and todos_eba:
        eba_combinado = pd.concat(todos_eba, ignore_index=True)
        
        # Ordenar por RP y después por NOMBRE ASEGURADO
        eba_combinado = eba_combinado.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
        
        print(f"Total registros EBA combinados: {len(eba_combinado)}")
    
    # Generar nombre del archivo
    mes_formateado = str(primer_mes).zfill(2)
    nombre_archivo = f"{mes_formateado}-{primer_anio}_MULTI_EMISION.xlsx"
    
    # Usar carpeta de destino si se especifica, sino usar folder_path
    if output_folder:
        base_path = output_folder
    else:
        base_path = folder_path
        
    output_path = os.path.join(base_path, nombre_archivo)
    
    try:
        # Crear el directorio si no existe
        os.makedirs(base_path, exist_ok=True)
        
        # Crear el archivo Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Escribir DataFrame ema_combinado en la hoja "EMA"
            ema_combinado.to_excel(writer, sheet_name='EMA', index=False)
            
            # Si existe eba_combinado, escribirlo en la hoja "EBA"
            if eba_combinado is not None:
                eba_combinado.to_excel(writer, sheet_name='EBA', index=False)
        
        # Aplicar formato a las hojas
        wb = load_workbook(output_path)
        
        # Formato para hoja EMA
        ws_ema = wb['EMA']
        header_fill = PatternFill(start_color='015d4d', end_color='015d4d', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Aplicar formato a los encabezados de EMA
        for cell in ws_ema[1]:  # Primera fila (encabezados)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columnas para EMA
        for column in ws_ema.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_ema.column_dimensions[column_letter].width = adjusted_width
        
        # Si existe hoja EBA, aplicar el mismo formato
        if eba_combinado is not None:
            ws_eba = wb['EBA']
            
            # Aplicar formato a los encabezados de EBA
            for cell in ws_eba[1]:  # Primera fila (encabezados)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas para EBA
            for column in ws_eba.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_eba.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar el archivo con formato
        wb.save(output_path)
        
        print(f"\n=== RESUMEN ===")
        print(f"Archivo Excel creado exitosamente: {output_path}")
        print(f"Archivos de emisión procesados: {len(archivos_emision_paths)}")
        print(f"Total registros EMA procesados: {len(ema_combinado)}")
        if eba_combinado is not None:
            print(f"Total registros EBA procesados: {len(eba_combinado)}")
        print("Formato aplicado: fondo #015d4d, fuente blanca")
        
        return output_path
        
    except Exception as e:
        print(f"Error al crear el archivo Excel: {e}")
        return None


def estructurar_1emision(emision_path, output_folder=None):
    """
    Función modificada para procesar un archivo de emisión individual.
    
    Args:
        emision_path (str): Ruta del archivo de emisión (.xls)
        output_folder (str): Carpeta donde guardar el resultado. Si es None, usa la carpeta del archivo original.
    
    Returns:
        str: Ruta del archivo Excel creado, o None si ocurre un error
    """
    
    # Cargar el archivo de Excel (.xls) usando xlrd
    wb = xlrd.open_workbook(emision_path)
    ws = wb.sheet_by_index(0)  # Primera hoja (índice 0)

    # Verificar el periodo en la celda B8 (fila 7, columna 1 en xlrd - índices basados en 0)
    periodo = ws.cell_value(7, 1)  # B8 = fila 7, columna 1
    if isinstance(periodo, str):
        mes, anio = periodo.split('/')
        mes = int(mes)
    else:
        raise ValueError("El formato de la celda B8 no es válido.")
    
    # Obtener el valor de RP de la celda B9 (fila 8, columna 1 en xlrd - índices basados en 0)
    rp_value = ws.cell_value(8, 1)  # B9 = fila 8, columna 1
    
    # Leer la hoja 2 (Emision Mensual)
    ema = pd.read_excel(emision_path, sheet_name=1, header=4, engine='xlrd')

    # Eliminar filas con el valor 2 en "Tipo del Movimiento"
    ema = ema[ema['Tipo del Movimiento'] != 2]
    # Sustituir todos los "#" por "Ñ" y eliminar espacios extra al final en la columna "Nombre"
    ema['Nombre'] = ema['Nombre'].str.replace('#', 'Ñ', regex=False).str.rstrip()
    # Eliminar las columnas "Origen del Movimiento" y "Tipo del Movimiento"
    ema = ema.drop(columns=['Origen del Movimiento', 'Tipo del Movimiento'], errors='ignore')
    # Cambiar el NSS a string para respetar los que tienen ceros al inicio
    ema['NSS'] = ema['NSS'].apply(lambda x: str(int(x)).zfill(11) if pd.notna(x) else x)
    # Limpiar y convertir "Fecha del Movimiento" a datetime
    ema['Fecha del Movimiento'] = ema['Fecha del Movimiento'].replace('-', pd.NaT)
    ema['Fecha del Movimiento'] = pd.to_datetime(ema['Fecha del Movimiento'], errors='coerce')
    # Ordenar por NSS y Fecha del Movimiento
    ema = ema.sort_values(['NSS', 'Fecha del Movimiento'], na_position='last')

    ema = ema.groupby('NSS').agg({
        'Nombre': 'first',
        'Días': 'sum',
        'Salario Diario': 'last',
        'Cuota Fija': 'sum',
        'Excedente Patronal': 'sum',
        'Excedente Obrero': 'sum',
        'Prestaciones en Dinero Patronal': 'sum',
        'Prestaciones en Dinero Obrero': 'sum',
        'Gastos Médicos y Pensionados Patronal': 'sum',
        'Gastos Médicos y Pensionados Obrero': 'sum',
        'Riesgos de Trabajo': 'sum',
        'Invalidez y Vida Patronal': 'sum',
        'Invalidez y Vida Obrero': 'sum',
        'Guarderías y Prestaciones Sociales': 'sum',
        'Total': 'sum'
    }).reset_index()

    # Agregar la columna RP como primera columna
    ema.insert(0, 'RP', rp_value)

    # Cambiar nombre a las columnas
    ema.columns = [
        'RP', 'NSS', 'NOMBRE ASEGURADO', 'DIAS', 'SDI', 'CF',
        'EXC_PAT', 'EXC_OBR', 'PD_PAT', 'PD_OBR', 'GMP_PAT',
        'GMP_OBR', 'RT', 'IV_PAT', 'IV_OBR', 'GPS', 'TOTAL'
    ]

    # Ordenar el DataFrame por RP y después por NOMBRE ASEGURADO
    ema = ema.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)

    # Si el mes es par, procesar también la hoja 3
    if mes % 2 == 0:
        print(f"Mes par ({mes}) detectado, procesando también la hoja 3...")
        
        # Leer la hoja 3
        eba = pd.read_excel(emision_path, sheet_name=2, header=4, engine='xlrd')

        # Procesar EBA
        eba = eba[eba['Tipo del Movimiento'] != 2]
        eba['Nombre'] = eba['Nombre'].str.replace('#', 'Ñ', regex=False).str.rstrip()
        eba = eba.drop(columns=['Origen del Movimiento', 'Tipo del Movimiento'], errors='ignore')
        eba['NSS'] = eba['NSS'].apply(lambda x: str(int(x)).zfill(11) if pd.notna(x) else x)
        # Normalizar "Número de Crédito" a 10 dígitos
        eba['Número de Crédito'] = eba['Número de Crédito'].apply(
            lambda x: str(x)[-10:] if pd.notna(x) else x
        )
        eba['Fecha del Movimiento'] = eba['Fecha del Movimiento'].replace('-', pd.NaT)
        eba['Fecha del Movimiento'] = pd.to_datetime(eba['Fecha del Movimiento'], errors='coerce')
        eba = eba.sort_values(['NSS', 'Fecha del Movimiento'], na_position='last')

        eba = eba.groupby('NSS').agg({
            'Nombre': 'first',
            'Días': 'sum',
            'Salario Diario': 'last',
            'Retiro': 'sum',
            'Cesantía en Edad Avanzada y Vejez Patronal': 'sum',
            'Cesantía en Edad Avanzada y Vejez Obrero': 'sum',
            'Subtotal RCV': 'sum',
            'Aportación Patronal': 'sum',
            'Tipo de Descuento': 'last',
            'Valor de Descuento': 'last',
            'Número de Crédito': 'last',
            'Amortización': 'sum',
            'Subtotal Infonavit': 'sum',
            'Total': 'sum'
        }).reset_index()

        # Agregar la columna RP como primera columna
        eba.insert(0, 'RP', rp_value)

        # Cambiar nombre a las columnas
        eba.columns = [
            'RP', 'NSS', 'NOMBRE ASEGURADO', 'DIAS', 'SDI', 'RETIRO',
            'CEAV_PAT', 'CEAV_OBR', 'TOTAL_RCV', 'APORTACION_PAT', 'T_CREDITO',
            'V_CREDITO', 'N_CREDITO', 'AMORTIZACION', 'TOTAL_INF', 'TOTAL'
        ]

        # Ordenar el DataFrame por RP y después por NOMBRE ASEGURADO
        eba = eba.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
        
        print(f"DataFrame eba creado con {len(eba)} filas")
    else:
        print(f"Mes impar ({mes}) detectado, solo se procesó la hoja 2")

    # Crear el nombre del archivo
    mes_formateado = str(mes).zfill(2)  # Agregar cero al inicio si es necesario
    nombre_archivo = f"{mes_formateado}-{anio}_{rp_value}_EMISION.xlsx"
    
    # Usar carpeta de destino si se especifica, sino usar carpeta original
    if output_folder:
        base_path = output_folder
    else:
        base_path = os.path.dirname(emision_path)
        
    output_path = os.path.join(base_path, nombre_archivo)
    
    # Crear el directorio si no existe
    os.makedirs(base_path, exist_ok=True)
    
    # Crear el archivo Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Escribir DataFrame ema en la hoja "EMA"
        ema.to_excel(writer, sheet_name='EMA', index=False)
        
        # Si existe eba, escribirlo en la hoja "EBA"
        if mes % 2 == 0:
            eba.to_excel(writer, sheet_name='EBA', index=False)
    
    # Aplicar formato a las hojas
    wb = load_workbook(output_path)
    
    # Formato para hoja EMA
    ws_ema = wb['EMA']
    header_fill = PatternFill(start_color='015d4d', end_color='015d4d', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Aplicar formato a los encabezados de EMA
    for cell in ws_ema[1]:  # Primera fila (encabezados)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar ancho de columnas para EMA
    for column in ws_ema.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_ema.column_dimensions[column_letter].width = adjusted_width
    
    # Si existe hoja EBA, aplicar el mismo formato
    if mes % 2 == 0:
        ws_eba = wb['EBA']
        
        # Aplicar formato a los encabezados de EBA
        for cell in ws_eba[1]:  # Primera fila (encabezados)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columnas para EBA
        for column in ws_eba.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_eba.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar el archivo con formato
    wb.save(output_path)
    
    print(f"Archivo Excel creado: {output_path}")
    return output_path
