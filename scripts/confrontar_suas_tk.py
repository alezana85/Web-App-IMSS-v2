import polars as pl
import pandas as pd
import os
import re
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

class ConfrontaSUAsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Confronta entre SUAs - IMSS")
        self.root.geometry("600x400")
        self.root.configure(bg='#f0f0f0')
        
        # Variables para las rutas de archivos
        self.sua1_path = tk.StringVar()
        self.sua2_path = tk.StringVar()
        self.output_path = tk.StringVar()
        
        self.setup_ui()
    
    def setup_ui(self):
        # Título principal
        title_label = tk.Label(self.root, text="Confronta entre Archivos SUA", 
                              font=("Arial", 16, "bold"), bg='#f0f0f0', fg='#333333')
        title_label.pack(pady=20)
        
        # Frame principal
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(padx=20, pady=10, fill='both', expand=True)
        
        # Sección para primer archivo SUA
        self.create_file_section(main_frame, "Primer archivo SUA:", self.sua1_path, 
                                self.select_sua1_file, 0)
        
        # Sección para segundo archivo SUA
        self.create_file_section(main_frame, "Segundo archivo SUA:", self.sua2_path, 
                                self.select_sua2_file, 1)
        
        # Sección para carpeta de destino
        self.create_file_section(main_frame, "Carpeta de destino:", self.output_path, 
                                self.select_output_folder, 2, is_folder=True)
        
        # Botón de procesar
        process_btn = tk.Button(main_frame, text="Procesar Confronta", 
                               command=self.process_files, bg='#4CAF50', fg='white',
                               font=("Arial", 12, "bold"), pady=10, cursor='hand2')
        process_btn.grid(row=3, column=0, columnspan=3, pady=30, sticky='ew')
        
        # Barra de progreso
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=4, column=0, columnspan=3, sticky='ew', pady=10)
        self.progress.grid_remove()  # Ocultar inicialmente
        
        # Label de estado
        self.status_label = tk.Label(main_frame, text="Listo para procesar", 
                                   bg='#f0f0f0', fg='#666666')
        self.status_label.grid(row=5, column=0, columnspan=3, pady=10)
        
        # Configurar pesos de columnas
        main_frame.columnconfigure(1, weight=1)
    
    def create_file_section(self, parent, label_text, var, command, row, is_folder=False):
        # Label
        label = tk.Label(parent, text=label_text, bg='#f0f0f0', 
                        font=("Arial", 10, "bold"))
        label.grid(row=row, column=0, sticky='w', pady=10)
        
        # Entry
        entry = tk.Entry(parent, textvariable=var, font=("Arial", 9), 
                        state='readonly', bg='white')
        entry.grid(row=row, column=1, sticky='ew', padx=10, pady=10)
        
        # Button
        btn_text = "Seleccionar Carpeta" if is_folder else "Seleccionar Archivo"
        button = tk.Button(parent, text=btn_text, command=command, 
                          bg='#2196F3', fg='white', cursor='hand2')
        button.grid(row=row, column=2, padx=5, pady=10)
    
    def select_sua1_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar primer archivo SUA",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.sua1_path.set(filename)
    
    def select_sua2_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar segundo archivo SUA",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.sua2_path.set(filename)
    
    def select_output_folder(self):
        folder = filedialog.askdirectory(title="Seleccionar carpeta de destino")
        if folder:
            self.output_path.set(folder)
    
    def validate_inputs(self):
        if not self.sua1_path.get():
            messagebox.showerror("Error", "Debe seleccionar el primer archivo SUA")
            return False
        if not self.sua2_path.get():
            messagebox.showerror("Error", "Debe seleccionar el segundo archivo SUA")
            return False
        if not self.output_path.get():
            messagebox.showerror("Error", "Debe seleccionar la carpeta de destino")
            return False
        
        # Verificar que los archivos existen
        if not os.path.exists(self.sua1_path.get()):
            messagebox.showerror("Error", "El primer archivo SUA no existe")
            return False
        if not os.path.exists(self.sua2_path.get()):
            messagebox.showerror("Error", "El segundo archivo SUA no existe")
            return False
        if not os.path.exists(self.output_path.get()):
            messagebox.showerror("Error", "La carpeta de destino no existe")
            return False
        
        return True
    
    def process_files(self):
        if not self.validate_inputs():
            return
        
        # Ejecutar en un hilo separado para no bloquear la UI
        thread = threading.Thread(target=self.run_confronta)
        thread.daemon = True
        thread.start()
    
    def run_confronta(self):
        try:
            # Actualizar UI en el hilo principal
            self.root.after(0, self.update_status, "Procesando archivos...")
            self.root.after(0, self.show_progress)
            
            # Ejecutar la confronta
            result_path = confronta_entre_suas(
                self.sua1_path.get(),
                self.sua2_path.get(),
                self.output_path.get()
            )
            
            # Actualizar UI con éxito
            self.root.after(0, self.hide_progress)
            self.root.after(0, self.update_status, f"¡Proceso completado exitosamente!")
            self.root.after(0, lambda: messagebox.showinfo(
                "Éxito", 
                f"Confronta completada exitosamente.\n\nArchivo generado:\n{result_path}"
            ))
            
        except Exception as e:
            # Actualizar UI con error
            self.root.after(0, self.hide_progress)
            self.root.after(0, self.update_status, "Error en el procesamiento")
            self.root.after(0, lambda: messagebox.showerror(
                "Error", 
                f"Error al procesar los archivos:\n\n{str(e)}"
            ))
    
    def update_status(self, message):
        self.status_label.config(text=message)
    
    def show_progress(self):
        self.progress.grid()
        self.progress.start()
    
    def hide_progress(self):
        self.progress.stop()
        self.progress.grid_remove()

def confronta_entre_suas(sua1_path, sua2_path, archive_path):
    sua1_filename = os.path.basename(sua1_path)
    sua2_filename = os.path.basename(sua2_path)
    
    sua1_match = re.search(r'(\d{2})-(\d{4})', sua1_filename)
    sua2_match = re.search(r'(\d{2})-(\d{4})', sua2_filename)
    
    if not sua1_match or not sua2_match:
        raise ValueError("Los nombres de archivo deben contener el formato MM-YYYY")
    
    sua1_mes, sua1_año = sua1_match.groups()
    sua2_mes, sua2_año = sua2_match.groups()
    
    mes = int(sua1_mes)
    año = int(sua1_año)
    
    num_hojas = 2 if mes % 2 == 0 else 1
    
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)
    
    purple_fill = PatternFill(start_color="8B008B", end_color="8B008B", fill_type="solid")
    green_font = Font(color="00FF00", bold=True)
    
    hoja1_result = procesar_hoja_sua_mensual(sua1_path, sua2_path)
    ws1 = wb.create_sheet("SUA_MENSUAL")
    escribir_dataframe_a_excel(ws1, hoja1_result, purple_fill, green_font)
    
    if num_hojas == 2:
        hoja2_result = procesar_hoja_sua_bimestral(sua1_path, sua2_path)
        ws2 = wb.create_sheet("SUA_BIMESTRAL")
        escribir_dataframe_a_excel(ws2, hoja2_result, purple_fill, green_font)
    
    output_filename = f"{sua1_mes}_{sua1_año}_CONFRONTA_SUAS.xlsx"
    output_path = os.path.join(archive_path, output_filename)
    wb.save(output_path)
    
    return output_path

def procesar_hoja_sua_mensual(sua1_path, sua2_path):
    try:
        sua1_df = pl.read_excel(sua1_path, sheet_name="SUA_MENSUAL")
        sua2_df = pl.read_excel(sua2_path, sheet_name="SUA_MENSUAL")
    except Exception as e:
        raise ValueError(f"Error al leer los archivos Excel: {e}")
    
    sua1_df = sua1_df.with_columns([
        pl.concat_str([pl.col("RP").cast(pl.Utf8).str.slice(0, 10), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    sua2_df = sua2_df.with_columns([
        pl.concat_str([pl.col("RP").cast(pl.Utf8).str.slice(0, 10), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    todas_ids = set(sua1_df["ID_UNICO"].to_list()) | set(sua2_df["ID_UNICO"].to_list())
    
    resultados = []
    
    for id_unico in todas_ids:
        sua1_row = sua1_df.filter(pl.col("ID_UNICO") == id_unico)
        sua2_row = sua2_df.filter(pl.col("ID_UNICO") == id_unico)
        
        sua1_existe = len(sua1_row) > 0
        sua2_existe = len(sua2_row) > 0
        
        if sua1_existe:
            sua1_data = sua1_row.to_dicts()[0]
            resultado = sua1_data.copy()
        else:
            sua2_data = sua2_row.to_dicts()[0]
            resultado = {col: None for col in sua1_df.columns}
            resultado["RP"] = sua2_data["RP"]
            resultado["NSS"] = sua2_data["NSS"]
            resultado["ID_UNICO"] = sua2_data["ID_UNICO"]
        
        if sua1_existe and sua2_existe:
            sua1_data = sua1_row.to_dicts()[0]
            sua2_data = sua2_row.to_dicts()[0]
            
            resultado["NOMBRE ASEGURADO"] = sua2_data["NOMBRE ASEGURADO"]
            resultado["RFC"] = sua2_data["RFC"]
            resultado["CURP"] = sua2_data["CURP"]
            resultado["N_MOVS"] = sua2_data["N_MOVS"]
            resultado["SDI"] = sua2_data["SDI"]
            resultado["INC"] = sua1_data.get("INC", 0)
            resultado["AUS"] = sua1_data.get("AUS", 0)
            
            columnas_numericas = ["DIAS", "CF", "EXC_PAT", "EXC_OBR", "PD_PAT", "PD_OBR", 
                                "GMP_PAT", "GMP_OBR", "RT", "IV_PAT", "IV_OBR", "GPS", "TOTAL"]
            
            for columna in columnas_numericas:
                if columna in sua1_data and columna in sua2_data:
                    sua1_val = sua1_data[columna] or 0
                    sua2_val = sua2_data[columna] or 0
                    
                    if isinstance(sua1_val, (float, int)) and isinstance(sua2_val, (float, int)):
                        sua1_val = round(float(sua1_val), 2)
                        sua2_val = round(float(sua2_val), 2)
                    
                    diferencia = sua1_val - sua2_val
                    diferencia = round(diferencia, 2)
                    resultado[columna] = diferencia
                    
        elif sua1_existe and not sua2_existe:
            pass
        elif not sua1_existe and sua2_existe:
            sua2_data = sua2_row.to_dicts()[0]
            resultado["NOMBRE ASEGURADO"] = sua2_data.get("NOMBRE ASEGURADO", "")
            for col in ["RFC", "CURP", "N_MOVS", "INC", "AUS"]:
                resultado[col] = None
        
        resultados.append(resultado)
    
    result_df = pl.DataFrame(resultados)
    return result_df.to_pandas()

def procesar_hoja_sua_bimestral(sua1_path, sua2_path):
    try:
        sua1_df = pl.read_excel(sua1_path, sheet_name="SUA_BIMESTRAL")
        sua2_df = pl.read_excel(sua2_path, sheet_name="SUA_BIMESTRAL")
    except Exception as e:
        raise ValueError(f"Error al leer las hojas SUA_BIMESTRAL: {e}")
    
    sua1_df = sua1_df.with_columns([
        pl.concat_str([pl.col("RP").cast(pl.Utf8).str.slice(0, 10), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    sua2_df = sua2_df.with_columns([
        pl.concat_str([pl.col("RP").cast(pl.Utf8).str.slice(0, 10), pl.col("NSS")], separator="_").alias("ID_UNICO")
    ])
    
    todas_ids = set(sua1_df["ID_UNICO"].to_list()) | set(sua2_df["ID_UNICO"].to_list())
    
    resultados = []
    
    for id_unico in todas_ids:
        sua1_row = sua1_df.filter(pl.col("ID_UNICO") == id_unico)
        sua2_row = sua2_df.filter(pl.col("ID_UNICO") == id_unico)
        
        sua1_existe = len(sua1_row) > 0
        sua2_existe = len(sua2_row) > 0
        
        if sua1_existe:
            sua1_data = sua1_row.to_dicts()[0]
            resultado = sua1_data.copy()
        else:
            sua2_data = sua2_row.to_dicts()[0]
            resultado = {col: None for col in sua1_df.columns}
            resultado["RP"] = sua2_data["RP"]
            resultado["NSS"] = sua2_data["NSS"]
            resultado["ID_UNICO"] = sua2_data["ID_UNICO"]
        
        if sua1_existe and sua2_existe:
            sua1_data = sua1_row.to_dicts()[0]
            sua2_data = sua2_row.to_dicts()[0]
            
            resultado["NOMBRE ASEGURADO"] = sua2_data["NOMBRE ASEGURADO"]
            resultado["RFC"] = sua2_data["RFC"]
            resultado["CURP"] = sua2_data["CURP"]
            resultado["N_MOVS"] = sua2_data["N_MOVS"]
            resultado["SDI"] = sua2_data["SDI"]
            resultado["N_CREDITO"] = sua2_data["N_CREDITO"]
            resultado["INC"] = sua1_data.get("INC", 0)
            resultado["AUS"] = sua1_data.get("AUS", 0)
            
            columnas_numericas = ["DIAS", "RETIRO", "CEAV_PAT", "CEAV_OBR", "TOTAL_RCV", 
                                "APORTACION_PAT", "AMORTIZACION", "TOTAL_INF", "TOTAL"]
            
            for columna in columnas_numericas:
                if columna in sua1_data and columna in sua2_data:
                    sua1_val = sua1_data[columna] or 0
                    sua2_val = sua2_data[columna] or 0
                    
                    if isinstance(sua1_val, (float, int)) and isinstance(sua2_val, (float, int)):
                        sua1_val = round(float(sua1_val), 2)
                        sua2_val = round(float(sua2_val), 2)
                    
                    diferencia = sua1_val - sua2_val
                    diferencia = round(diferencia, 2)
                    resultado[columna] = diferencia
                    
        elif sua1_existe and not sua2_existe:
            pass
        elif not sua1_existe and sua2_existe:
            sua2_data = sua2_row.to_dicts()[0]
            resultado["NOMBRE ASEGURADO"] = sua2_data.get("NOMBRE ASEGURADO", "")
            for col in ["RFC", "CURP", "N_MOVS", "N_CREDITO", "INC", "AUS"]:
                resultado[col] = None
        
        resultados.append(resultado)
    
    result_df = pl.DataFrame(resultados)
    return result_df.to_pandas()

def escribir_dataframe_a_excel(worksheet, df, header_fill, header_font):
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
    
    for row_num, row_data in enumerate(df.itertuples(index=False), 2):
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    root = tk.Tk()
    app = ConfrontaSUAsApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
