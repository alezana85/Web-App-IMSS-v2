import os
import pandas as pd
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def estructurar_1emision(emision_path):
    # Cargar el archivo de Excel (.xls) usando xlrd
    wb = xlrd.open_workbook(emision_path)
    ws = wb.sheet_by_index(0)  # Primera hoja (índice 0)

    """
    Tenemos que verificar la celda B8 de la primera hoja del arcihivo para ver a que periodo pertenece
    el cual esta en el formato mm/yyyy, esto determinara si procederemos a ller la hoja 2 y 3 o solo la 2,
    si el perido es mes par entonces se leeran al hojas 2 y 3, si es impar solo la hoja 2.

    El perido si es un mes de un digito no tiene el 0 al inicio, por ejemplo: 1/2023
    """
    # Verificar el periodo en la celda B8 (fila 7, columna 1 en xlrd - índices basados en 0)
    periodo = ws.cell_value(7, 1)  # B8 = fila 7, columna 1
    if isinstance(periodo, str):
        mes, anio = periodo.split('/')
        mes = int(mes)
    else:
        raise ValueError("El formato de la celda B8 no es válido.")
    
    # Obtener el valor de RP de la celda B9 (fila 8, columna 1 en xlrd - índices basados en 0)
    rp_value = ws.cell_value(8, 1)  # B9 = fila 8, columna 1
    
    # Leer la hoja 2 (Emision Mensual)
    # El encabezado de la hoja 2 esta en la fila 5 y para manejar mejor los datos tenemos que convertirlos a csv
    ema = pd.read_excel(emision_path, sheet_name=1, header=4, engine='xlrd')

    # Eliminar filas con el valor 2 en "Tipo del Movimiento"
    ema = ema[ema['Tipo del Movimiento'] != 2]
    # Sustituir todos los "#" por "Ñ" y eliminar espacios extra al final en la columna "Nombre"
    ema['Nombre'] = ema['Nombre'].str.replace('#', 'Ñ', regex=False).str.rstrip()
    # Eliminar las columnas "Origen del Movimiento" y "Tipo del Movimiento"
    ema = ema.drop(columns=['Origen del Movimiento', 'Tipo del Movimiento'], errors='ignore')
    # Cambiar el NSS a string para respetar los que tienen ceros al inicio
    ema['NSS'] = ema['NSS'].apply(lambda x: str(int(x)).zfill(11) if pd.notna(x) else x)
    # Limpiar y convertir "Fecha del Movimiento" a datetime
    # Reemplazar "-" con NaN antes de convertir a datetime
    ema['Fecha del Movimiento'] = ema['Fecha del Movimiento'].replace('-', pd.NaT)
    ema['Fecha del Movimiento'] = pd.to_datetime(ema['Fecha del Movimiento'], errors='coerce')
    # Eliminar filas completamente duplicadas
    ema = ema.drop_duplicates()
    # Ordenar por NSS y Fecha del Movimiento para asegurar que 'last' tome el más reciente
    # Los valores NaT (fechas faltantes) irán al final del ordenamiento
    ema = ema.sort_values(['NSS', 'Fecha del Movimiento'], na_position='last')

    ema = ema.groupby('NSS').agg({
        'Nombre': 'first',  # Mantener el primer nombre encontrado
        'Días': 'sum',
        'Salario Diario': 'last',  # Tomar el último salario diario según la fecha del movimiento
        'Cuota Fija': 'sum',
        'Excedente Patronal': 'sum',
        'Excedente Obrero': 'sum',
        'Prestaciones en Dinero Patronal': 'sum',
        'Prestaciones en Dinero Obrero': 'sum',
        'Gastos Médicos y Pensionados Patronal': 'sum',
        'Gastos Médicos y Pensionados Obrero': 'sum',
        'Riesgos de Trabajo': 'sum',
        'Invalidez y Vida Patronal': 'sum',
        'Invalidez y Vida Obrero': 'sum',
        'Guarderías y Prestaciones Sociales': 'sum',
        'Total': 'sum'
    }).reset_index()

    # Agregar la columna RP como primera columna con el valor obtenido de B9
    ema.insert(0, 'RP', rp_value)

    # Cambiar nombre a las columnas
    ema.columns = [
        'RP', 'NSS', 'NOMBRE ASEGURADO', 'DIAS', 'SDI', 'CF',
        'EXC_PAT', 'EXC_OBR', 'PD_PAT', 'PD_OBR', 'GMP_PAT',
        'GMP_OBR', 'RT', 'IV_PAT', 'IV_OBR', 'GPS', 'TOTAL'
    ]

    # Ordenar el DataFrame por RP y después por NOMBRE ASEGURADO
    ema = ema.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)

    # Si el mes es par, procesar también la hoja 3
    if mes % 2 == 0:
        print(f"Mes par ({mes}) detectado, procesando también la hoja 3...")
        
        # Leer la hoja 3 (similar a la hoja 2)
        eba = pd.read_excel(emision_path, sheet_name=2, header=4, engine='xlrd')

        # Eliminar filas con el valor 2 en "Tipo del Movimiento"
        eba = eba[eba['Tipo del Movimiento'] != 2]
        # Sustituir todos los "#" por "Ñ" y eliminar espacios extra al final en la columna "Nombre"
        eba['Nombre'] = eba['Nombre'].str.replace('#', 'Ñ', regex=False).str.rstrip()
        # Eliminar las columnas "Origen del Movimiento" y "Tipo del Movimiento"
        eba = eba.drop(columns=['Origen del Movimiento', 'Tipo del Movimiento'], errors='ignore')
        # Cambiar el NSS a string para respetar los que tienen ceros al inicio
        eba['NSS'] = eba['NSS'].apply(lambda x: str(int(x)).zfill(11) if pd.notna(x) else x)
        # Normalizar "Número de Crédito" a 10 dígitos (eliminando ceros extra a la izquierda si tiene 11)
        eba['Número de Crédito'] = eba['Número de Crédito'].apply(
            lambda x: str(x)[-10:] if pd.notna(x) else x
        )
        # Limpiar y convertir "Fecha del Movimiento" a datetime
        # Reemplazar "-" con NaN antes de convertir a datetime
        eba['Fecha del Movimiento'] = eba['Fecha del Movimiento'].replace('-', pd.NaT)
        eba['Fecha del Movimiento'] = pd.to_datetime(eba['Fecha del Movimiento'], errors='coerce')
        # Eliminar filas completamente duplicadas
        eba = eba.drop_duplicates()
        # Ordenar por NSS y Fecha del Movimiento para asegurar que 'last' tome el más reciente
        # Los valores NaT (fechas faltantes) irán al final del ordenamiento
        eba = eba.sort_values(['NSS', 'Fecha del Movimiento'], na_position='last')

        eba = eba.groupby('NSS').agg({
            'Nombre': 'first',  # Mantener el primer nombre encontrado
            'Días': 'sum',
            'Salario Diario': 'last',  # Tomar el último salario diario según la fecha del movimiento
            'Retiro': 'sum',
            'Cesantía en Edad Avanzada y Vejez Patronal': 'sum',
            'Cesantía en Edad Avanzada y Vejez Obrero': 'sum',
            'Subtotal RCV': 'sum',
            'Aportación Patronal': 'sum',
            'Tipo de Descuento': 'last',
            'Valor de Descuento': 'last',
            'Número de Crédito': 'last',
            'Amortización': 'sum',
            'Subtotal Infonavit': 'sum',
            'Total': 'sum'
        }).reset_index()

        # Agregar la columna RP como primera columna con el valor obtenido de B9
        eba.insert(0, 'RP', rp_value)

        # Cambiar nombre a las columnas
        eba.columns = [
            'RP', 'NSS', 'NOMBRE ASEGURADO', 'DIAS', 'SDI', 'RETIRO',
            'CEAV_PAT', 'CEAV_OBR', 'TOTAL_RCV', 'APORTACION_PAT', 'T_CREDITO',
            'V_CREDITO', 'N_CREDITO', 'AMORTIZACION', 'TOTAL_INF', 'TOTAL'
        ]

        # Ordenar el DataFrame por RP y después por NOMBRE ASEGURADO
        eba = eba.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
        
        print(f"DataFrame eba creado con {len(eba)} filas")
    else:
        print(f"Mes impar ({mes}) detectado, solo se procesó la hoja 2")

    # Crear archivo Excel con openpyxl
    # Crear el nombre del archivo
    mes_formateado = str(mes).zfill(2)  # Agregar cero al inicio si es necesario
    nombre_archivo = f"{mes_formateado}-{anio}_{rp_value}_EMISION.xlsx"
    
    # Crear el directorio de salida si no existe
    output_dir = os.path.dirname(emision_path)
    output_path = os.path.join(output_dir, nombre_archivo)
    
    # Crear el archivo Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Escribir DataFrame ema en la hoja "EMA"
        ema.to_excel(writer, sheet_name='EMA', index=False)
        
        # Si existe eba, escribirlo en la hoja "EBA"
        if mes % 2 == 0:
            eba.to_excel(writer, sheet_name='EBA', index=False)
    
    # Aplicar formato a las hojas
    wb = load_workbook(output_path)
    
    # Formato para hoja EMA
    ws_ema = wb['EMA']
    header_fill = PatternFill(start_color='015d4d', end_color='015d4d', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Aplicar formato a los encabezados de EMA
    for cell in ws_ema[1]:  # Primera fila (encabezados)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar ancho de columnas para EMA
    for column in ws_ema.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_ema.column_dimensions[column_letter].width = adjusted_width
    
    # Si existe hoja EBA, aplicar el mismo formato
    if mes % 2 == 0:
        ws_eba = wb['EBA']
        
        # Aplicar formato a los encabezados de EBA
        for cell in ws_eba[1]:  # Primera fila (encabezados)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columnas para EBA
        for column in ws_eba.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_eba.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar el archivo con formato
    wb.save(output_path)
    
    print(f"Archivo Excel creado: {output_path}")

def estrucurar_varias_emisiones(folder_path):
    """
    Función para procesar múltiples archivos de emisión en una carpeta y combinarlos 
    en un solo archivo Excel con hojas EMA y EBA (si aplica).
    
    Args:
        folder_path (str): Ruta de la carpeta que contiene los archivos de emisión
    
    Returns:
        str: Ruta del archivo Excel creado, o None si ocurre un error
    """
    
    if not os.path.exists(folder_path):
        print(f"La carpeta {folder_path} no existe.")
        return None
    
    def buscar_archivos_emision(directorio, nivel=0, max_nivel=5):
        """Busca archivos .xls de emisión recursivamente hasta el nivel máximo especificado"""
        archivos_emision = []
        
        if nivel > max_nivel:
            return archivos_emision
        
        try:
            for item in os.listdir(directorio):
                item_path = os.path.join(directorio, item)
                
                if os.path.isfile(item_path) and item.endswith('.xls'):
                    # Verificar que sea un archivo de emisión válido
                    try:
                        wb = xlrd.open_workbook(item_path)
                        ws = wb.sheet_by_index(0)
                        # Verificar si tiene el formato esperado (periodo en B8)
                        periodo = ws.cell_value(7, 1)  # B8
                        if isinstance(periodo, str) and '/' in periodo:
                            archivos_emision.append(item_path)
                    except Exception:
                        # Si no puede leer el archivo o no tiene el formato esperado, lo omite
                        continue
                elif os.path.isdir(item_path) and nivel < max_nivel:
                    archivos_emision.extend(buscar_archivos_emision(item_path, nivel + 1, max_nivel))
        except PermissionError:
            print(f"Sin permisos para acceder a: {directorio}")
        except Exception as e:
            print(f"Error accediendo a {directorio}: {e}")
        
        return archivos_emision
    
    # Buscar todos los archivos de emisión en la carpeta y subcarpetas
    print("Buscando archivos de emisión (.xls) en carpeta y subcarpetas (máximo 5 niveles)...")
    archivos_emision_paths = buscar_archivos_emision(folder_path)
    
    if not archivos_emision_paths:
        print("No se encontraron archivos de emisión válidos en la carpeta especificada ni en subcarpetas.")
        return None
    
    print(f"Se encontraron {len(archivos_emision_paths)} archivos de emisión para procesar:")
    for archivo_path in archivos_emision_paths:
        print(f"  - {archivo_path}")
    
    # Listas para almacenar todos los registros
    todos_ema = []
    todos_eba = []
    
    # Variables para determinar el nombre del archivo
    primer_mes = None
    primer_anio = None
    crear_hoja_eba = False
    
    # Procesar cada archivo de emisión
    for archivo_path in archivos_emision_paths:
        archivo_nombre = os.path.basename(archivo_path)
        print(f"Procesando archivo: {archivo_nombre}")
        
        try:
            # Cargar el archivo de Excel (.xls) usando xlrd
            wb = xlrd.open_workbook(archivo_path)
            ws = wb.sheet_by_index(0)  # Primera hoja
            
            # Verificar el periodo en la celda B8
            periodo = ws.cell_value(7, 1)  # B8 = fila 7, columna 1
            if isinstance(periodo, str):
                mes, anio = periodo.split('/')
                mes = int(mes)
                anio = int(anio)
            else:
                print(f"Formato de periodo inválido en {archivo_nombre}, omitiendo archivo.")
                continue
            
            # Usar el primer archivo para determinar el nombre del archivo final
            if primer_mes is None:
                primer_mes = mes
                primer_anio = anio
                crear_hoja_eba = (mes % 2 == 0)
            
            # Obtener el valor de RP de la celda B9
            rp_value = ws.cell_value(8, 1)  # B9 = fila 8, columna 1
            
            # Procesar hoja 2 (Emisión Mensual - EMA)
            try:
                ema = pd.read_excel(archivo_path, sheet_name=1, header=4, engine='xlrd')
                
                # Aplicar las mismas transformaciones que en la función individual
                ema = ema[ema['Tipo del Movimiento'] != 2]
                ema['Nombre'] = ema['Nombre'].str.replace('#', 'Ñ', regex=False).str.rstrip()
                ema = ema.drop(columns=['Origen del Movimiento', 'Tipo del Movimiento'], errors='ignore')
                ema['NSS'] = ema['NSS'].apply(lambda x: str(int(x)).zfill(11) if pd.notna(x) else x)
                ema['Fecha del Movimiento'] = ema['Fecha del Movimiento'].replace('-', pd.NaT)
                ema['Fecha del Movimiento'] = pd.to_datetime(ema['Fecha del Movimiento'], errors='coerce')
                # Eliminar filas completamente duplicadas
                ema = ema.drop_duplicates()
                ema = ema.sort_values(['NSS', 'Fecha del Movimiento'], na_position='last')
                
                ema = ema.groupby('NSS').agg({
                    'Nombre': 'first',
                    'Días': 'sum',
                    'Salario Diario': 'last',
                    'Cuota Fija': 'sum',
                    'Excedente Patronal': 'sum',
                    'Excedente Obrero': 'sum',
                    'Prestaciones en Dinero Patronal': 'sum',
                    'Prestaciones en Dinero Obrero': 'sum',
                    'Gastos Médicos y Pensionados Patronal': 'sum',
                    'Gastos Médicos y Pensionados Obrero': 'sum',
                    'Riesgos de Trabajo': 'sum',
                    'Invalidez y Vida Patronal': 'sum',
                    'Invalidez y Vida Obrero': 'sum',
                    'Guarderías y Prestaciones Sociales': 'sum',
                    'Total': 'sum'
                }).reset_index()
                
                # Agregar la columna RP como primera columna con el valor obtenido de B9
                ema.insert(0, 'RP', rp_value)
                
                # Cambiar nombres de columnas
                ema.columns = [
                    'RP', 'NSS', 'NOMBRE ASEGURADO', 'DIAS', 'SDI', 'CF',
                    'EXC_PAT', 'EXC_OBR', 'PD_PAT', 'PD_OBR', 'GMP_PAT',
                    'GMP_OBR', 'RT', 'IV_PAT', 'IV_OBR', 'GPS', 'TOTAL'
                ]
                
                todos_ema.append(ema)
                print(f"Procesado {archivo_nombre} - EMA: {len(ema)} registros")
                
            except Exception as e:
                print(f"Error procesando hoja EMA en {archivo_nombre}: {e}")
                continue
            
            # Procesar hoja 3 (Emisión Bimestral - EBA) solo si el mes es par
            if crear_hoja_eba:
                try:
                    eba = pd.read_excel(archivo_path, sheet_name=2, header=4, engine='xlrd')
                    
                    # Aplicar las mismas transformaciones que en la función individual
                    eba = eba[eba['Tipo del Movimiento'] != 2]
                    eba['Nombre'] = eba['Nombre'].str.replace('#', 'Ñ', regex=False).str.rstrip()
                    eba = eba.drop(columns=['Origen del Movimiento', 'Tipo del Movimiento'], errors='ignore')
                    eba['NSS'] = eba['NSS'].apply(lambda x: str(int(x)).zfill(11) if pd.notna(x) else x)
                    eba['Número de Crédito'] = eba['Número de Crédito'].apply(
                        lambda x: str(x)[-10:] if pd.notna(x) else x
                    )
                    eba['Fecha del Movimiento'] = eba['Fecha del Movimiento'].replace('-', pd.NaT)
                    eba['Fecha del Movimiento'] = pd.to_datetime(eba['Fecha del Movimiento'], errors='coerce')
                    # Eliminar filas completamente duplicadas
                    eba = eba.drop_duplicates()
                    eba = eba.sort_values(['NSS', 'Fecha del Movimiento'], na_position='last')
                    
                    eba = eba.groupby('NSS').agg({
                        'Nombre': 'first',
                        'Días': 'sum',
                        'Salario Diario': 'last',
                        'Retiro': 'sum',
                        'Cesantía en Edad Avanzada y Vejez Patronal': 'sum',
                        'Cesantía en Edad Avanzada y Vejez Obrero': 'sum',
                        'Subtotal RCV': 'sum',
                        'Aportación Patronal': 'sum',
                        'Tipo de Descuento': 'last',
                        'Valor de Descuento': 'last',
                        'Número de Crédito': 'last',
                        'Amortización': 'sum',
                        'Subtotal Infonavit': 'sum',
                        'Total': 'sum'
                    }).reset_index()
                    
                    # Agregar la columna RP como primera columna con el valor obtenido de B9
                    eba.insert(0, 'RP', rp_value)
                    
                    # Cambiar nombres de columnas
                    eba.columns = [
                        'RP', 'NSS', 'NOMBRE ASEGURADO', 'DIAS', 'SDI', 'RETIRO',
                        'CEAV_PAT', 'CEAV_OBR', 'TOTAL_RCV', 'APORTACION_PAT', 'T_CREDITO',
                        'V_CREDITO', 'N_CREDITO', 'AMORTIZACION', 'TOTAL_INF', 'TOTAL'
                    ]
                    
                    todos_eba.append(eba)
                    print(f"Procesado {archivo_nombre} - EBA: {len(eba)} registros")
                    
                except Exception as e:
                    print(f"Error procesando hoja EBA en {archivo_nombre}: {e}")
                    continue
            
        except Exception as e:
            print(f"Error procesando archivo {archivo_nombre}: {e}")
            continue
    
    # Verificar que se encontraron registros
    if not todos_ema:
        print("No se encontraron registros EMA válidos en ningún archivo.")
        return None
    
    # Combinar todos los DataFrames EMA
    ema_combinado = pd.concat(todos_ema, ignore_index=True)
    
    # Ordenar por RP y después por NOMBRE ASEGURADO
    ema_combinado = ema_combinado.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
    
    print(f"Total registros EMA combinados: {len(ema_combinado)}")
    
    # Combinar todos los DataFrames EBA si existen
    eba_combinado = None
    if crear_hoja_eba and todos_eba:
        eba_combinado = pd.concat(todos_eba, ignore_index=True)
        
        # Ordenar por RP y después por NOMBRE ASEGURADO
        eba_combinado = eba_combinado.sort_values(['RP', 'NOMBRE ASEGURADO']).reset_index(drop=True)
        
        print(f"Total registros EBA combinados: {len(eba_combinado)}")
    
    # Generar nombre del archivo
    mes_formateado = str(primer_mes).zfill(2)
    nombre_archivo = f"{mes_formateado}-{primer_anio}_MULTI_EMISION.xlsx"
    output_path = os.path.join(folder_path, nombre_archivo)
    
    try:
        # Crear el archivo Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Escribir DataFrame EMA combinado
            ema_combinado.to_excel(writer, sheet_name='EMA', index=False)
            
            # Escribir DataFrame EBA combinado si existe
            if eba_combinado is not None:
                eba_combinado.to_excel(writer, sheet_name='EBA', index=False)
        
        # Aplicar formato a las hojas
        wb = load_workbook(output_path)
        
        # Formato para hoja EMA
        ws_ema = wb['EMA']
        header_fill = PatternFill(start_color='015d4d', end_color='015d4d', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Aplicar formato a los encabezados de EMA
        for cell in ws_ema[1]:  # Primera fila (encabezados)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar ancho de columnas para EMA
        for column in ws_ema.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_ema.column_dimensions[column_letter].width = adjusted_width
        
        # Si existe hoja EBA, aplicar el mismo formato
        if eba_combinado is not None:
            ws_eba = wb['EBA']
            
            # Aplicar formato a los encabezados de EBA
            for cell in ws_eba[1]:  # Primera fila (encabezados)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas para EBA
            for column in ws_eba.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_eba.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar el archivo con formato
        wb.save(output_path)
        
        print(f"\n=== RESUMEN ===")
        print(f"Archivo Excel creado exitosamente: {output_path}")
        print(f"Archivos de emisión procesados: {len(archivos_emision_paths)}")
        print(f"Total registros EMA procesados: {len(ema_combinado)}")
        if eba_combinado is not None:
            print(f"Total registros EBA procesados: {len(eba_combinado)}")
            print("Hoja EBA incluida (mes par detectado)")
        print("Formato aplicado: fondo #015d4d, fuente blanca")
        
        return output_path
        
    except Exception as e:
        print(f"Error al crear el archivo Excel: {e}")
        return None